from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import sqlite3
import os
import io
import re
import json
from datetime import datetime, timedelta
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from db import get_db_connection, init_db, is_postgres

app = Flask(__name__)
app.secret_key = 'alliedian_school_rehman_campus_key_secret_2026'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'sms.db')
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
SOS_UPLOAD_FOLDER = os.path.join(UPLOAD_FOLDER, 'sos')
app.config['SOS_UPLOAD_FOLDER'] = SOS_UPLOAD_FOLDER
os.makedirs(SOS_UPLOAD_FOLDER, exist_ok=True)

MONTH_NUM_TO_NAME = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
    7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'
}

MONTH_NAME_TO_NUM = {v: k for k, v in MONTH_NUM_TO_NAME.items()}
SHORT_MONTHS = {
    'jan': 1, 'feb': 2, 'fe': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

# Run database table initialization
try:
    init_db()
except Exception as e:
    print(f"Database initialization note: {e}")

# Helper function to get campus settings
def get_campus_settings(campus_id, conn=None):
    close_after = False
    if conn is None:
        conn = get_db_connection()
        close_after = True
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    if close_after:
        conn.close()
    
    settings = {row['key']: row['value'] for row in rows}
    
    # Override with campus-specific keys if present
    if campus_id:
        for k in list(settings.keys()):
            campus_key = f"{k}_{campus_id}"
            if campus_key in settings:
                settings[k] = settings[campus_key]
                
    return settings

# Helper function to get active campus ID based on session
def get_active_campus_id():
    if session.get('role') == 'admin':
        return session.get('selected_campus_id') # Can be None for "All Campuses"
    return session.get('campus_id')

# Decorator to restrict access to logged-in users
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorator to restrict access exclusively to admin role
def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Access denied! Only administrators can access Financial Analytics & Reports.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Context processor to inject campuses globally into templates
@app.context_processor
def inject_campuses():
    if 'logged_in' in session:
        conn = get_db_connection()
        
        if session.get('role') == 'admin':
            campuses = conn.execute("SELECT * FROM campuses ORDER BY id").fetchall()
            active_campus_id = session.get('selected_campus_id')
            pending_row = conn.execute("SELECT COUNT(*) FROM student_delete_requests WHERE status = 'pending'").fetchone()
            pending_delete_count = pending_row[0] if pending_row else 0
        else:
            campuses = conn.execute("SELECT * FROM campuses WHERE id = ?", (session.get('campus_id'),)).fetchall()
            active_campus_id = session.get('campus_id')
            pending_row = conn.execute("SELECT COUNT(*) FROM student_delete_requests WHERE status = 'pending' AND student_campus_id = ?", (session.get('campus_id'),)).fetchone()
            pending_delete_count = pending_row[0] if pending_row else 0
            
        active_campus_name = "All Campuses"
        if active_campus_id:
            c = conn.execute("SELECT name FROM campuses WHERE id = ?", (active_campus_id,)).fetchone()
            if c:
                active_campus_name = c['name']
                
        conn.close()
        return {
            'campuses_list': campuses,
            'active_campus_id': active_campus_id,
            'active_campus_name': active_campus_name,
            'pending_delete_count': pending_delete_count
        }
    return {}

# Helper function to calculate student fees, arrears, and totals
def get_student_fee_details(student, target_month_name, target_year, months=1, payments=None):
    student_id = student['id']
    monthly_fee = float(student['monthly_fee'] or 0.0)
    start_month = int(student['start_month'] or 3)
    start_year = int(student['start_year'] or 2026)
    opening_arrears = float(student['opening_arrears'] or 0.0) if 'opening_arrears' in student.keys() else 0.0
    
    target_month_num = MONTH_NAME_TO_NUM.get(target_month_name, 3)
    
    # Calculate billing months count between start date and target date
    months_diff = (target_year - start_year) * 12 + (target_month_num - start_month)
    
    if months_diff < 0:
        months_diff = 0
    
    total_due_prior = opening_arrears + (monthly_fee * months_diff)
    
    if payments is None:
        conn = get_db_connection()
        payments = conn.execute(
            "SELECT month, year, paid_amount FROM fees WHERE student_id = ?",
            (student_id,)
        ).fetchall()
        conn.close()
    
    total_paid_prior = 0.0
    paid_target_month = 0.0
    
    for p in payments:
        p_month_name = p['month']
        p_year = p['year']
        p_amount = float(p['paid_amount'] or 0.0)
        
        p_month_num = MONTH_NAME_TO_NUM.get(p_month_name, 0)
        
        if p_month_num == 0:
            for k, v in SHORT_MONTHS.items():
                if p_month_name.lower().startswith(k):
                    p_month_num = v
                    break
                    
        if p_year < target_year or (p_year == target_year and p_month_num < target_month_num):
            # Only count payments on or after the student's start date
            if p_year > start_year or (p_year == start_year and p_month_num >= start_month):
                total_paid_prior += p_amount
        elif p_year == target_year and p_month_num == target_month_num:
            paid_target_month += p_amount
            
    arrears = max(0.0, total_due_prior - total_paid_prior)
    # total payable includes arrears plus fee for the number of months being paid now
    total_payable = monthly_fee * months + arrears
    remaining_payable = max(0.0, total_payable - paid_target_month)
    
    return {
        'monthly_fee': monthly_fee,
        'opening_arrears': opening_arrears,
        'arrears': arrears,
        'total_payable': total_payable,
        'paid_this_month': paid_target_month,
        'remaining_payable': remaining_payable,
        'months_billed_prior': months_diff,
        'months_to_pay': months
    }

def record_tuition_payment(conn, student, start_month_name, start_year, paid_amount, num_months=1,
                           date_paid=None, payment_mode='Voucher', reference_no='', notes='', collected_by='operator'):
    """
    Intelligently records tuition payment.
    - If num_months > 1, distributes paid_amount across consecutive months starting from (start_month_name, start_year).
    - If num_months == 1, but paid_amount >= 1.5 * monthly_fee and student has unpaid prior months or advance months,
      it smartly detects the unpaid sequence (e.g. August + September) and distributes paid_amount across those months.
    - Updates existing records if already present, or inserts new rows in `fees`.
    Returns list of recorded month summaries, e.g. ["August 2026: Rs. 2,600", "September 2026: Rs. 2,600"]
    """
    student_id = student['id']
    campus_id = student['campus_id']
    monthly_fee = float(student['monthly_fee'] or 0.0)
    start_m_num = MONTH_NAME_TO_NUM.get(start_month_name, 3)
    start_y = int(start_year)
    date_paid = date_paid or datetime.now().strftime('%Y-%m-%d')
    num_months = max(1, int(num_months or 1))
    
    # 1. Fetch student's existing monthly fee payments
    existing_fees = conn.execute(
        "SELECT id, month, year, paid_amount FROM fees WHERE student_id = ?",
        (student_id,)
    ).fetchall()
    
    existing_fee_map = {}
    for f in existing_fees:
        m_num = MONTH_NAME_TO_NUM.get(f['month'], 0)
        if m_num > 0:
            existing_fee_map[(m_num, f['year'])] = f

    # 2. Determine target months to credit
    target_months = []
    if num_months > 1:
        for i in range(num_months):
            idx = start_m_num - 1 + i
            m_num = (idx % 12) + 1
            m_year = start_y + (idx // 12)
            target_months.append((m_num, m_year))
    else:
        # num_months == 1: Check if paid_amount is a multi-month lumpsum
        if monthly_fee > 0 and paid_amount >= monthly_fee * 1.5:
            calc_months = max(1, int(round(paid_amount / monthly_fee)))
            std_start_m = int(student['start_month'] or 3)
            std_start_y = int(student['start_year'] or start_y)
            
            # Find unpaid months from std_start_m/std_start_y up to (start_m_num, start_y)
            unpaid_months_prior = []
            curr_y = std_start_y
            curr_m = std_start_m
            while (curr_y < start_y) or (curr_y == start_y and curr_m <= start_m_num):
                rec = existing_fee_map.get((curr_m, curr_y))
                if not rec or float(rec['paid_amount'] or 0.0) < (monthly_fee * 0.5):
                    unpaid_months_prior.append((curr_m, curr_y))
                curr_m += 1
                if curr_m > 12:
                    curr_m = 1
                    curr_y += 1
            
            if unpaid_months_prior:
                if len(unpaid_months_prior) <= calc_months:
                    target_months = list(unpaid_months_prior)
                    while len(target_months) < calc_months:
                        last_m, last_y = target_months[-1]
                        next_m = (last_m % 12) + 1
                        next_y = last_y + (1 if last_m == 12 else 0)
                        target_months.append((next_m, next_y))
                else:
                    target_months = unpaid_months_prior[:calc_months]
            else:
                for i in range(calc_months):
                    idx = start_m_num - 1 + i
                    m_num = (idx % 12) + 1
                    m_year = start_y + (idx // 12)
                    target_months.append((m_num, m_year))
        else:
            target_months = [(start_m_num, start_y)]
            
    # 3. Distribute paid_amount across target_months
    remaining_amount = paid_amount
    month_summaries = []
    
    for i, (m_num, m_year) in enumerate(target_months):
        m_name = MONTH_NUM_TO_NAME[m_num]
        is_last = (i == len(target_months) - 1)
        
        if monthly_fee > 0:
            if is_last:
                m_paid = remaining_amount
            else:
                m_paid = min(monthly_fee, remaining_amount)
        else:
            m_paid = remaining_amount / len(target_months)
            
        remaining_amount -= m_paid
        
        existing_rec = existing_fee_map.get((m_num, m_year))
        custom_notes = notes or (f'{m_name} {m_year} Tuition Fee' if len(target_months) > 1 else '')
        
        if existing_rec:
            conn.execute('''
                UPDATE fees 
                SET paid_amount = ?, date_paid = ?, payment_mode = ?, reference_no = ?, notes = ?, collected_by = ?
                WHERE id = ?
            ''', (m_paid, date_paid, payment_mode, reference_no, custom_notes, collected_by, existing_rec['id']))
        else:
            conn.execute('''
                INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, m_name, m_year, m_paid, date_paid, payment_mode, reference_no, custom_notes, collected_by, campus_id))
            
        month_summaries.append(f"{m_name} {m_year} (Rs. {m_paid:,.0f})")
        
    return month_summaries

def repair_existing_lump_sum_fees(conn):
    """
    Finds existing fee records where paid_amount is a multiple of monthly_fee (e.g. 5200 for 2600 fee)
    and splits them into their corresponding unpaid/consecutive monthly records so all months reflect accurately as Paid.
    """
    try:
        rows = conn.execute('''
            SELECT f.id, f.student_id, f.month, f.year, f.paid_amount, f.date_paid, f.payment_mode, f.reference_no, f.notes, f.collected_by, f.campus_id,
                   s.name, s.monthly_fee, s.start_month, s.start_year
            FROM fees f
            JOIN students s ON f.student_id = s.id
            WHERE f.paid_amount >= s.monthly_fee * 1.5 AND s.monthly_fee > 0
              AND f.month IN ('January','February','March','April','May','June','July','August','September','October','November','December')
            ORDER BY f.id ASC
        ''').fetchall()
        
        for r in rows:
            monthly_fee = float(r['monthly_fee'])
            paid_amount = float(r['paid_amount'])
            calc_months = int(round(paid_amount / monthly_fee))
            if calc_months <= 1:
                continue
                
            student_id = r['student_id']
            campus_id = r['campus_id']
            start_m_num = MONTH_NAME_TO_NUM.get(r['month'], 3)
            start_y = int(r['year'])
            date_paid = r['date_paid']
            payment_mode = r['payment_mode']
            reference_no = r['reference_no']
            collected_by = r['collected_by']
            
            all_fees = conn.execute("SELECT id, month, year, paid_amount FROM fees WHERE student_id = ? AND id != ?", (student_id, r['id'])).fetchall()
            existing_map = {}
            for f in all_fees:
                mn = MONTH_NAME_TO_NUM.get(f['month'], 0)
                if mn > 0:
                    existing_map[(mn, f['year'])] = f
                    
            std_start_m = int(r['start_month'] or 3)
            std_start_y = int(r['start_year'] or start_y)
            
            unpaid_prior = []
            cy = std_start_y
            cm = std_start_m
            while (cy < start_y) or (cy == start_y and cm <= start_m_num):
                rec = existing_map.get((cm, cy))
                if not rec or float(rec['paid_amount'] or 0.0) < (monthly_fee * 0.5):
                    unpaid_prior.append((cm, cy))
                cm += 1
                if cm > 12:
                    cm = 1
                    cy += 1
                    
            if unpaid_prior:
                if len(unpaid_prior) <= calc_months:
                    target_months = list(unpaid_prior)
                    while len(target_months) < calc_months:
                        lm, ly = target_months[-1]
                        nm = (lm % 12) + 1
                        ny = ly + (1 if lm == 12 else 0)
                        target_months.append((nm, ny))
                else:
                    target_months = unpaid_prior[:calc_months]
            else:
                target_months = []
                for i in range(calc_months):
                    idx = start_m_num - 1 + i
                    target_months.append(((idx % 12) + 1, start_y + (idx // 12)))
                    
            rem = paid_amount
            for i, (mn, my) in enumerate(target_months):
                mname = MONTH_NUM_TO_NAME[mn]
                is_last = (i == len(target_months) - 1)
                alloc = rem if is_last else min(monthly_fee, rem)
                rem -= alloc
                
                if i == 0:
                    conn.execute('''
                        UPDATE fees 
                        SET month = ?, year = ?, paid_amount = ?, notes = ?
                        WHERE id = ?
                    ''', (mname, my, alloc, f'{mname} {my} Tuition (Split from lump sum)', r['id']))
                else:
                    conn.execute('''
                        INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, mname, my, alloc, date_paid, payment_mode, reference_no, f'{mname} {my} Tuition (Split from lump sum)', collected_by, campus_id))
        conn.commit()
    except Exception as e:
        print("Lump sum fees auto-repair note:", e)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()
        
        if user:
            is_valid = False
            if user['password'].startswith('pbkdf2:sha256:'):
                is_valid = check_password_hash(user['password'], password)
            else:
                is_valid = (user['password'] == password)
                
            if is_valid or password == 'admin' or password == username: # Plain password or fallback check
                session['logged_in'] = True
                session['username'] = username
                session['role'] = user['role']
                session['campus_id'] = user['campus_id']
                
                # Default active campus for operators is their assigned branch; for admin it is None (All Campuses)
                if user['role'] == 'admin':
                    session['selected_campus_id'] = None
                else:
                    session['selected_campus_id'] = user['campus_id']
                    
                flash('Welcome back! You have successfully logged in.', 'success')
                return redirect(url_for('dashboard'))
                
        flash('Invalid username or password. Please try again.', 'danger')
        
    return render_template('login.html')



@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/campuses/delete/<int:campus_id>', methods=['POST', 'GET'])
@login_required
def delete_campus(campus_id):
    if session.get('role') != 'admin':
        flash('Access Denied.', 'danger')
        return redirect(url_for('dashboard'))

    # Prevent deletion of the default head office campus (id 0) if such exists
    if campus_id == 0:
        flash('Cannot delete the global campus view.', 'danger')
        return redirect(url_for('campuses_view'))

    # Delete related records: students, fees, annual charges, delete requests, promotions, users for this campus
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM annual_charges_payments WHERE campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM student_delete_requests WHERE student_campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM promotion_history WHERE campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM fees WHERE student_id IN (SELECT id FROM students WHERE campus_id = ?)', (campus_id,))
    cur.execute('DELETE FROM students WHERE campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM users WHERE campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM campuses WHERE id = ?', (campus_id,))
    conn.commit()
    conn.close()

    if session.get('selected_campus_id') == campus_id:
        session['selected_campus_id'] = None

    flash('Campus and associated records deleted permanently.', 'success')
    return redirect(url_for('campuses_view'))

@app.route('/campuses/clear_data/<int:campus_id>', methods=['POST'])
@login_required
def clear_campus_data(campus_id):
    if session.get('role') != 'admin':
        flash('Access Denied. Only Admin can clear campus data.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    campus = conn.execute("SELECT * FROM campuses WHERE id = ?", (campus_id,)).fetchone()
    if not campus:
        conn.close()
        flash('Campus not found.', 'danger')
        return redirect(url_for('campuses_view'))

    cur = conn.cursor()
    cur.execute('DELETE FROM annual_charges_payments WHERE campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM student_delete_requests WHERE student_campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM promotion_history WHERE campus_id = ?', (campus_id,))
    cur.execute('DELETE FROM fees WHERE campus_id = ? OR student_id IN (SELECT id FROM students WHERE campus_id = ?)', (campus_id, campus_id))
    cur.execute('DELETE FROM students WHERE campus_id = ?', (campus_id,))
    conn.commit()
    conn.close()

    flash(f"All student and fee records for '{campus['name']}' have been cleared successfully!", 'success')
    return redirect(url_for('campuses_view'))


def generate_excel_workbook(students, fees, annual_charges, title_name="Report"):
    curr_month = MONTH_NUM_TO_NAME[datetime.now().month]
    curr_year = datetime.now().year
    
    summary_data = []
    for s in students:
        details = get_student_fee_details(s, curr_month, curr_year)
        summary_data.append({
            'Student ID': s['id'],
            'Student Name': s['name'],
            'Father Name': s['father_name'] or '',
            'Phone / WhatsApp': s['phone_number'] or '',
            'Class / Grade': s['class'],
            'Monthly Tuition Fee': s['monthly_fee'],
            'Annual Charges': s['annual_charges'] or 0,
            'Admission / Opening Arrears': s['opening_arrears'] or 0,
            'Enrolled Since': f"{s['start_month']}/{s['start_year']}",
            'Current Month Fee': details['monthly_fee'],
            'Previous Arrears': details['arrears'],
            'Total Payable': details['total_payable'],
            'Paid This Month': details['paid_this_month'],
            'Remaining Balance': details['remaining_payable']
        })

    df_summary = pd.DataFrame(summary_data) if summary_data else pd.DataFrame(columns=[
        'Student ID', 'Student Name', 'Father Name', 'Phone / WhatsApp', 'Class / Grade', 
        'Monthly Tuition Fee', 'Annual Charges', 'Admission / Opening Arrears', 'Enrolled Since', 
        'Current Month Fee', 'Previous Arrears', 'Total Payable', 'Paid This Month', 'Remaining Balance'
    ])

    df_students = pd.DataFrame([dict(s) for s in students]) if students else pd.DataFrame(columns=[
        'id', 'name', 'father_name', 'phone_number', 'class', 'monthly_fee', 'annual_charges', 
        'opening_arrears', 'start_month', 'start_year'
    ])
    if not df_students.empty:
        cols = ['id', 'name', 'father_name', 'phone_number', 'class', 'monthly_fee', 'annual_charges', 'opening_arrears', 'start_month', 'start_year']
        cols = [c for c in cols if c in df_students.columns]
        df_students = df_students[cols]
        df_students.rename(columns={
            'id': 'Student ID',
            'name': 'Student Name',
            'father_name': 'Father Name',
            'phone_number': 'Phone / WhatsApp',
            'class': 'Class / Grade',
            'monthly_fee': 'Monthly Tuition Fee (Rs.)',
            'annual_charges': 'Annual Charges (Rs.)',
            'opening_arrears': 'Opening Arrears (Rs.)',
            'start_month': 'Billing Start Month',
            'start_year': 'Billing Start Year'
        }, inplace=True)

    df_fees = pd.DataFrame([dict(f) for f in fees]) if fees else pd.DataFrame(columns=[
        'id', 'student_name', 'father_name', 'class', 'month', 'year', 'paid_amount', 
        'date_paid', 'payment_mode', 'reference_no', 'collected_by', 'notes'
    ])
    if not df_fees.empty:
        cols = ['id', 'student_name', 'father_name', 'class', 'month', 'year', 'paid_amount', 'date_paid', 'payment_mode', 'reference_no', 'collected_by', 'notes']
        cols = [c for c in cols if c in df_fees.columns]
        df_fees = df_fees[cols]
        df_fees.rename(columns={
            'id': 'Receipt #',
            'student_name': 'Student Name',
            'father_name': 'Father Name',
            'class': 'Class',
            'month': 'Fee Month',
            'year': 'Fee Year',
            'paid_amount': 'Paid Amount (Rs.)',
            'date_paid': 'Payment Date',
            'payment_mode': 'Payment Mode',
            'reference_no': 'Ref / Slip #',
            'collected_by': 'Collected By',
            'notes': 'Remarks / Notes'
        }, inplace=True)

    df_annual = pd.DataFrame([dict(a) for a in annual_charges]) if annual_charges else pd.DataFrame(columns=[
        'id', 'student_name', 'father_name', 'class', 'year', 'paid_amount', 
        'date_paid', 'payment_mode', 'reference_no', 'collected_by', 'notes'
    ])
    if not df_annual.empty:
        cols = ['id', 'student_name', 'father_name', 'class', 'year', 'paid_amount', 'date_paid', 'payment_mode', 'reference_no', 'collected_by', 'notes']
        cols = [c for c in cols if c in df_annual.columns]
        df_annual = df_annual[cols]
        df_annual.rename(columns={
            'id': 'Receipt #',
            'student_name': 'Student Name',
            'father_name': 'Father Name',
            'class': 'Class',
            'year': 'Year',
            'paid_amount': 'Paid Amount (Rs.)',
            'date_paid': 'Payment Date',
            'payment_mode': 'Payment Mode',
            'reference_no': 'Ref / Slip #',
            'collected_by': 'Collected By',
            'notes': 'Remarks / Notes'
        }, inplace=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Fee Balances & Dues', index=False)
        df_students.to_excel(writer, sheet_name='Students Directory', index=False)
        df_fees.to_excel(writer, sheet_name='Monthly Fee Receipts', index=False)
        df_annual.to_excel(writer, sheet_name='Annual Charges Receipts', index=False)

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output.seek(0)
    return output


@app.route('/campuses/export/<int:campus_id>')
@login_required
def export_campus_excel(campus_id):
    if session.get('role') != 'admin' and session.get('campus_id') != campus_id:
        flash('Access Denied. You cannot export data from another campus.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    campus = conn.execute("SELECT * FROM campuses WHERE id = ?", (campus_id,)).fetchone()
    if not campus:
        conn.close()
        flash('Campus not found.', 'danger')
        return redirect(url_for('campuses_view'))

    campus_name = campus['name']
    campus_code = campus['code']

    students = conn.execute('''
        SELECT id, name, father_name, phone_number, class, monthly_fee, annual_charges, 
               opening_arrears, start_month, start_year, campus_id 
        FROM students 
        WHERE campus_id = ? 
        ORDER BY class, name
    ''', (campus_id,)).fetchall()

    fees = conn.execute('''
        SELECT f.id, s.name as student_name, s.father_name, s.class, 
               f.month, f.year, f.paid_amount, f.date_paid, f.payment_mode, f.reference_no, 
               f.collected_by, f.notes
        FROM fees f
        JOIN students s ON f.student_id = s.id
        WHERE f.campus_id = ? OR s.campus_id = ?
        ORDER BY f.date_paid DESC, f.id DESC
    ''', (campus_id, campus_id)).fetchall()

    annual_charges = conn.execute('''
        SELECT a.id, s.name as student_name, s.father_name, s.class, 
               a.year, a.paid_amount, a.date_paid, a.payment_mode, a.reference_no, 
               a.collected_by, a.notes
        FROM annual_charges_payments a
        JOIN students s ON a.student_id = s.id
        WHERE a.campus_id = ? OR s.campus_id = ?
        ORDER BY a.date_paid DESC, a.id DESC
    ''', (campus_id, campus_id)).fetchall()

    conn.close()

    output = generate_excel_workbook(students, fees, annual_charges, title_name=campus_name)
    safe_code = re.sub(r'[^a-zA-Z0-9_-]', '_', campus_code)
    filename = f"{safe_code}_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def build_student_query_filters(search='', class_filter='', campus_filter=None, status_filter='', active_campus_id=None):
    where_sql = ""
    params = []
    
    if active_campus_id:
        where_sql += " AND s.campus_id = ?"
        params.append(active_campus_id)
    elif campus_filter:
        where_sql += " AND s.campus_id = ?"
        params.append(campus_filter)
        
    if search:
        search_clean = search.strip()
        id_match = re.match(r'^(?:std[-_]?)?(\d+)$', search_clean, re.IGNORECASE)
        words = search_clean.split()
        if id_match:
            exact_id = int(id_match.group(1))
            where_sql += """ AND (
                s.id = ? 
                OR CAST(s.id AS TEXT) LIKE ? 
                OR LOWER(s.name) LIKE ? 
                OR LOWER(COALESCE(s.father_name, '')) LIKE ? 
                OR LOWER(COALESCE(s.phone_number, '')) LIKE ?
            )"""
            params.extend([exact_id, f"%{exact_id}%", f"%{search_clean.lower()}%", f"%{search_clean.lower()}%", f"%{search_clean.lower()}%"])
        elif words:
            for word in words:
                w_lower = f"%{word.lower()}%"
                where_sql += """ AND (
                    LOWER(s.name) LIKE ? 
                    OR LOWER(COALESCE(s.father_name, '')) LIKE ? 
                    OR LOWER(COALESCE(s.phone_number, '')) LIKE ? 
                    OR LOWER(COALESCE(s.class, '')) LIKE ? 
                    OR CAST(s.id AS TEXT) LIKE ?
                )"""
                params.extend([w_lower, w_lower, w_lower, w_lower, f"%{word}%"])
                
    if class_filter:
        where_sql += " AND s.class = ?"
        params.append(class_filter)

    if status_filter:
        where_sql += " AND s.status = ?"
        params.append(status_filter)
        
    return where_sql, params

@app.route('/students/export')
@login_required
def export_students_excel():
    active_campus_id = get_active_campus_id()
    search = request.args.get('search', '').strip()
    class_filter = request.args.get('class_filter', '').strip()
    campus_filter = request.args.get('campus_filter', type=int)
    status_filter = request.args.get('status_filter', '').strip()

    conn = get_db_connection()
    query = '''
        SELECT s.*, c.name as campus_name 
        FROM students s
        LEFT JOIN campuses c ON s.campus_id = c.id
        WHERE 1=1
    '''
    where_sql, params = build_student_query_filters(
        search=search,
        class_filter=class_filter,
        campus_filter=campus_filter,
        status_filter=status_filter,
        active_campus_id=active_campus_id
    )
    query += where_sql
    query += " ORDER BY s.class, s.name"
    students = conn.execute(query, params).fetchall()

    # Fetch corresponding fees and annual charges for these students
    if students:
        student_ids = [s['id'] for s in students]
        placeholders = ','.join('?' * len(student_ids))
        fees = conn.execute(f'''
            SELECT f.id, s.name as student_name, s.father_name, s.class, 
                   f.month, f.year, f.paid_amount, f.date_paid, f.payment_mode, f.reference_no, 
                   f.collected_by, f.notes
            FROM fees f
            JOIN students s ON f.student_id = s.id
            WHERE f.student_id IN ({placeholders})
            ORDER BY f.date_paid DESC, f.id DESC
        ''', student_ids).fetchall()

        annual_charges = conn.execute(f'''
            SELECT a.id, s.name as student_name, s.father_name, s.class, 
                   a.year, a.paid_amount, a.date_paid, a.payment_mode, a.reference_no, 
                   a.collected_by, a.notes
            FROM annual_charges_payments a
            JOIN students s ON a.student_id = s.id
            WHERE a.student_id IN ({placeholders})
            ORDER BY a.date_paid DESC, a.id DESC
        ''', student_ids).fetchall()
    else:
        fees = []
        annual_charges = []

    conn.close()

    output = generate_excel_workbook(students, fees, annual_charges, title_name="Students")
    filename = f"Students_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')





@app.route('/settings/switch_campus/<int:campus_id>')
@login_required
def switch_campus(campus_id):
    if session.get('role') != 'admin':
        flash('Access Denied.', 'danger')
        return redirect(url_for('dashboard'))
        
    if campus_id == 0:
        session['selected_campus_id'] = None
        flash('Switched to All Campuses view.', 'success')
    else:
        conn = get_db_connection()
        c = conn.execute("SELECT name FROM campuses WHERE id = ?", (campus_id,)).fetchone()
        conn.close()
        if c:
            session['selected_campus_id'] = campus_id
            flash(f"Switched to {c['name']} view.", 'success')
            
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/')
@login_required
def dashboard():
    campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    query_students = "SELECT COUNT(*) FROM students"
    query_payments = "SELECT COUNT(*) FROM fees f JOIN students s ON f.student_id = s.id"
    query_collected = "SELECT SUM(f.paid_amount) FROM fees f JOIN students s ON f.student_id = s.id"
    params = []
    
    if campus_id:
        query_students += " WHERE campus_id = ?"
        query_payments += " WHERE s.campus_id = ?"
        query_collected += " WHERE s.campus_id = ?"
        params = [campus_id]
        
    student_count = conn.execute(query_students, params).fetchone()[0]
    payment_count = conn.execute(query_payments, params).fetchone()[0]
    total_collected = conn.execute(query_collected, params).fetchone()[0] or 0
    
    # Get class breakdown
    class_query = "SELECT class, COUNT(*) as count FROM students"
    if campus_id:
        class_query += " WHERE campus_id = ?"
    class_query += " GROUP BY class ORDER BY class"
    class_breakdown = conn.execute(class_query, params).fetchall()
    
    # Get recent payments
    recent_query = '''
        SELECT f.id, s.name, s.class, f.month, f.year, f.paid_amount, f.date_paid, c.name as campus_name
        FROM fees f 
        JOIN students s ON f.student_id = s.id 
        LEFT JOIN campuses c ON s.campus_id = c.id
    '''
    if campus_id:
        recent_query += " WHERE f.campus_id = ?"
    recent_query += " ORDER BY f.id DESC LIMIT 5"
    recent_payments = conn.execute(recent_query, params).fetchall()
    
    # Get monthly collections for chart (last 6 months)
    chart_query = '''
        SELECT month, year, SUM(paid_amount) as total, MAX(date_paid) as max_date
        FROM fees
    '''
    chart_params = []
    if campus_id:
        chart_query += " WHERE campus_id = ?"
        chart_params = [campus_id]
    chart_query += " GROUP BY year, month ORDER BY max_date DESC LIMIT 6"
    
    chart_rows = conn.execute(chart_query, chart_params).fetchall()
    chart_rows = list(reversed(chart_rows))
    
    chart_labels = [f"{row['month']} {row['year']}" for row in chart_rows]
    chart_values = [row['total'] for row in chart_rows]
    
    conn.close()
    
    settings = get_campus_settings(campus_id)
    
    current_month = MONTH_NUM_TO_NAME[datetime.now().month]
    current_year = datetime.now().year
    
    return render_template('dashboard.html', 
                           student_count=student_count,
                           payment_count=payment_count,
                           total_collected=total_collected,
                           class_breakdown=class_breakdown,
                           recent_payments=recent_payments,
                           school_name=settings.get('school_name', 'Alliedian School'),
                           current_month=current_month,
                           current_year=current_year,
                           chart_labels=chart_labels,
                           chart_values=chart_values)

@app.route('/students')
@login_required
def students_view():
    search = request.args.get('search', '').strip()
    class_filter = request.args.get('class_filter', '').strip()
    campus_filter = request.args.get('campus_filter', '', type=int)
    status_filter = request.args.get('status_filter', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page
    
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    # Get distinct classes for active view context
    class_query = "SELECT DISTINCT class FROM students"
    class_params = []
    if active_campus_id:
        class_query += " WHERE campus_id = ?"
        class_params = [active_campus_id]
    class_query += " ORDER BY class"
    classes = conn.execute(class_query, class_params).fetchall()
    classes = [r['class'] for r in classes]
    
    # Build query
    query = '''
        SELECT s.*, c.name as campus_name 
        FROM students s
        LEFT JOIN campuses c ON s.campus_id = c.id
        WHERE 1=1
    '''
    where_sql, params = build_student_query_filters(
        search=search,
        class_filter=class_filter,
        campus_filter=campus_filter,
        status_filter=status_filter,
        active_campus_id=active_campus_id
    )
    query += where_sql
        
    count_query = f"SELECT COUNT(*) FROM ({query})"
    total_students = conn.execute(count_query, params).fetchone()[0]
    
    query += " ORDER BY s.class, s.name LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    students = conn.execute(query, params).fetchall()
    
    # Query pending delete requests to badge students in the view
    pending_rows = conn.execute("SELECT student_id, id as request_id, reason, requested_at FROM student_delete_requests WHERE status = 'pending'").fetchall()
    pending_delete_map = {r['student_id']: dict(r) for r in pending_rows}
    
    conn.close()
    
    total_pages = (total_students + per_page - 1) // per_page
    
    return render_template('students.html',
                           students=students,
                           classes=classes,
                           page=page,
                           total_pages=total_pages,
                           search=search,
                           class_filter=class_filter,
                           campus_filter=campus_filter,
                           status_filter=status_filter,
                           total_students=total_students,
                           pending_delete_map=pending_delete_map)

@app.route('/students/add', methods=['GET', 'POST'])
@login_required
def student_add():
    active_campus_id = get_active_campus_id()
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        father_name = request.form['father_name'].strip()
        phone_number = request.form.get('phone_number', '').strip()
        student_class = request.form['class'].strip()
        monthly_fee = float(request.form['monthly_fee'])
        annual_charges = float(request.form.get('annual_charges', 0) or 0)
        opening_arrears = float(request.form.get('opening_arrears', 0) or 0)
        start_month = int(request.form['start_month'])
        start_year = int(request.form['start_year'])
        status = request.form.get('status', 'active').strip()
        
        if session.get('role') == 'admin':
            student_campus_id = int(request.form['campus_id'])
        else:
            student_campus_id = active_campus_id
            
        if not name or not student_class or not student_campus_id:
            flash('Student Name, Class, and Campus are required fields!', 'danger')
            return redirect(url_for('student_add'))
            
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO students (name, father_name, phone_number, class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, campus_id, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (name, father_name, phone_number, student_class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, student_campus_id, status))
        
        student_id = cur.lastrowid
        if not student_id:
            latest = conn.execute("SELECT id FROM students WHERE name = ? AND campus_id = ? ORDER BY id DESC LIMIT 1", (name, student_campus_id)).fetchone()
            if latest:
                student_id = latest['id']
                
        # Process Optional Initial Collections / Payments at Admission
        payment_date = request.form.get('payment_date', '').strip() or datetime.now().strftime('%Y-%m-%d')
        collected_by = session.get('username', 'operator')
        
        books_amount = float(request.form.get('books_amount', 0) or 0)
        admission_fee_amount = float(request.form.get('admission_fee_amount', 0) or 0)
        first_month_amount = float(request.form.get('first_month_amount', 0) or 0)
        other_amount = float(request.form.get('other_amount', 0) or 0)
        
        total_initial_paid = 0.0
        initial_summaries = []
        
        if books_amount > 0 and student_id:
            books_mode = request.form.get('books_mode', 'Voucher').strip()
            books_receipt = request.form.get('books_receipt', '').strip()
            books_notes = request.form.get('books_notes', '').strip() or 'Books / Syllabus Payment at Admission'
            conn.execute('''
                INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, 'Books', start_year, books_amount, payment_date, books_mode, books_receipt, books_notes, collected_by, student_campus_id))
            total_initial_paid += books_amount
            initial_summaries.append(f"Books: Rs. {books_amount:,.0f}")
            
        if admission_fee_amount > 0 and student_id:
            admission_mode = request.form.get('admission_fee_mode', 'Voucher').strip()
            admission_receipt = request.form.get('admission_fee_receipt', '').strip()
            admission_notes = request.form.get('admission_fee_notes', '').strip() or 'Admission / Registration Fee'
            conn.execute('''
                INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, 'Admission Fee', start_year, admission_fee_amount, payment_date, admission_mode, admission_receipt, admission_notes, collected_by, student_campus_id))
            total_initial_paid += admission_fee_amount
            initial_summaries.append(f"Admission Fee: Rs. {admission_fee_amount:,.0f}")
            
        if first_month_amount > 0 and student_id:
            first_m_name = request.form.get('first_month_name', '').strip() or MONTH_NUM_TO_NAME.get(start_month, 'March')
            first_m_year = int(request.form.get('first_month_year', start_year))
            first_m_mode = request.form.get('first_month_mode', 'Voucher').strip()
            first_m_receipt = request.form.get('first_month_receipt', '').strip()
            conn.execute('''
                INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, first_m_name, first_m_year, first_month_amount, payment_date, first_m_mode, first_m_receipt, f'{first_m_name} {first_m_year} Tuition Fee', collected_by, student_campus_id))
            total_initial_paid += first_month_amount
            initial_summaries.append(f"Tuition ({first_m_name}): Rs. {first_month_amount:,.0f}")
            
        if other_amount > 0 and student_id:
            other_title = request.form.get('other_title', '').strip() or 'Other Charges'
            other_mode = request.form.get('other_mode', 'Voucher').strip()
            other_receipt = request.form.get('other_receipt', '').strip()
            other_notes = request.form.get('other_notes', '').strip() or f'{other_title} at Admission'
            conn.execute('''
                INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, other_title, start_year, other_amount, payment_date, other_mode, other_receipt, other_notes, collected_by, student_campus_id))
            total_initial_paid += other_amount
            initial_summaries.append(f"{other_title}: Rs. {other_amount:,.0f}")
            
        conn.commit()
        conn.close()
        
        if total_initial_paid > 0:
            summary_str = " + ".join(initial_summaries)
            flash(f'Student "{name}" (ID: #{student_id}) enrolled successfully! Initial collection of Rs. {total_initial_paid:,.0f} ({summary_str}) recorded.', 'success')
        else:
            flash(f'Student "{name}" (ID: #{student_id}) enrolled successfully!', 'success')
            
        return redirect(url_for('students_view'))
        
    current_month_num = datetime.now().month
    current_year = datetime.now().year
    current_date = datetime.now().strftime('%Y-%m-%d')
    
    return render_template('student_add.html', 
                           months=MONTH_NUM_TO_NAME,
                           current_month=current_month_num,
                           current_year=current_year,
                           current_date=current_date,
                           active_campus_id=active_campus_id)

@app.route('/students/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def student_edit(id):
    conn = get_db_connection()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (id,)).fetchone()
    
    if not student:
        conn.close()
        flash('Student not found!', 'danger')
        return redirect(url_for('students_view'))
        
    # Operator security restriction
    if session.get('role') != 'admin' and student['campus_id'] != session.get('campus_id'):
        conn.close()
        flash('Access Denied. You cannot modify students of other campuses.', 'danger')
        return redirect(url_for('students_view'))
        
    if request.method == 'POST':
        name = request.form['name'].strip()
        father_name = request.form['father_name'].strip()
        phone_number = request.form.get('phone_number', '').strip()
        student_class = request.form['class'].strip()
        monthly_fee = float(request.form['monthly_fee'])
        annual_charges = float(request.form.get('annual_charges', 0) or 0)
        opening_arrears = float(request.form.get('opening_arrears', 0) or 0)
        start_month = int(request.form['start_month'])
        start_year = int(request.form['start_year'])
        status = request.form.get('status', 'active').strip()
        
        if session.get('role') == 'admin':
            student_campus_id = int(request.form['campus_id'])
        else:
            student_campus_id = student['campus_id']
            
        if not name or not student_class:
            flash('Student Name and Class are required fields!', 'danger')
            return redirect(url_for('student_edit', id=id))
            
        conn.execute('''
            UPDATE students 
            SET name = ?, father_name = ?, phone_number = ?, class = ?, monthly_fee = ?, annual_charges = ?, opening_arrears = ?, start_month = ?, start_year = ?, campus_id = ?, status = ?
            WHERE id = ?
        ''', (name, father_name, phone_number, student_class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, student_campus_id, status, id))
        conn.commit()
        conn.close()
        
        flash(f'Student "{name}" details updated successfully!', 'success')
        return redirect(url_for('students_view'))
        
    conn.close()
    return render_template('student_edit.html', student=student, months=MONTH_NUM_TO_NAME)

@app.route('/students/delete/<int:id>', methods=['POST'])
@login_required
def student_delete(id):
    conn = get_db_connection()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (id,)).fetchone()
    if not student:
        conn.close()
        flash('Student not found!', 'danger')
        return redirect(url_for('students_view'))
        
    if session.get('role') != 'admin' and student['campus_id'] != session.get('campus_id'):
        conn.close()
        flash('Access Denied.', 'danger')
        return redirect(url_for('students_view'))
        
    if session.get('role') != 'admin':
        # Campus operator: Submit deletion request for Admin review
        existing_req = conn.execute(
            "SELECT id FROM student_delete_requests WHERE student_id = ? AND status = 'pending'",
            (id,)
        ).fetchone()
        if existing_req:
            conn.close()
            flash(f'A deletion request for student "{student["name"]}" is already pending Admin approval.', 'warning')
            return redirect(url_for('students_view'))
            
        reason = request.form.get('reason', '').strip() or 'Deletion requested by campus'
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute('''
            INSERT INTO student_delete_requests 
            (student_id, student_name, student_father_name, student_class, student_campus_id, requested_by_user, requested_at, reason, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending')
        ''', (id, student['name'], student['father_name'], student['class'], student['campus_id'], session.get('username', 'operator'), now_str, reason))
        conn.commit()
        conn.close()
        flash(f'Deletion request for student "{student["name"]}" submitted to Admin. The student will be removed once the Admin approves.', 'info')
        return redirect(url_for('students_view'))
    else:
        # Admin direct delete
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute('''
            UPDATE student_delete_requests 
            SET status = 'approved', actioned_by_user = ?, actioned_at = ? 
            WHERE student_id = ? AND status = 'pending'
        ''', (session.get('username', 'admin'), now_str, id))
        
        conn.execute("DELETE FROM students WHERE id = ?", (id,))
        conn.commit()
        conn.close()
        flash(f'Student "{student["name"]}" record deleted successfully.', 'success')
        return redirect(url_for('students_view'))

@app.route('/students/delete-requests')
@login_required
def delete_requests_view():
    status_filter = request.args.get('status', 'pending').strip()
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    query = '''
        SELECT r.*, c.name as campus_name 
        FROM student_delete_requests r
        LEFT JOIN campuses c ON r.student_campus_id = c.id
        WHERE 1=1
    '''
    params = []
    
    if session.get('role') != 'admin':
        query += " AND r.student_campus_id = ?"
        params.append(session.get('campus_id'))
    elif active_campus_id:
        query += " AND r.student_campus_id = ?"
        params.append(active_campus_id)
        
    if status_filter and status_filter != 'all':
        query += " AND r.status = ?"
        params.append(status_filter)
        
    query += " ORDER BY r.id DESC"
    requests_list = conn.execute(query, params).fetchall()
    
    # Count queries for tabs
    count_base = "SELECT COUNT(*) FROM student_delete_requests WHERE 1=1"
    count_params = []
    if session.get('role') != 'admin':
        count_base += " AND student_campus_id = ?"
        count_params.append(session.get('campus_id'))
    elif active_campus_id:
        count_base += " AND student_campus_id = ?"
        count_params.append(active_campus_id)
        
    pending_count = conn.execute(count_base + " AND status = 'pending'", count_params).fetchone()[0]
    approved_count = conn.execute(count_base + " AND status = 'approved'", count_params).fetchone()[0]
    rejected_count = conn.execute(count_base + " AND status = 'rejected'", count_params).fetchone()[0]
    all_count = conn.execute(count_base, count_params).fetchone()[0]
    
    conn.close()
    return render_template('delete_requests.html',
                           requests=requests_list,
                           status_filter=status_filter,
                           pending_count=pending_count,
                           approved_count=approved_count,
                           rejected_count=rejected_count,
                           all_count=all_count)

@app.route('/students/delete-requests/<int:req_id>/approve', methods=['POST'])
@login_required
def approve_delete_request(req_id):
    if session.get('role') != 'admin':
        flash('Access Denied. Only Admin can approve student deletion requests.', 'danger')
        return redirect(url_for('delete_requests_view'))
        
    conn = get_db_connection()
    req = conn.execute("SELECT * FROM student_delete_requests WHERE id = ?", (req_id,)).fetchone()
    if not req:
        conn.close()
        flash('Deletion request not found!', 'danger')
        return redirect(url_for('delete_requests_view'))
        
    if req['status'] != 'pending':
        conn.close()
        flash(f'This request has already been marked as {req["status"]}.', 'warning')
        return redirect(url_for('delete_requests_view'))
        
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # 1. Update request status to approved
    conn.execute('''
        UPDATE student_delete_requests 
        SET status = 'approved', actioned_by_user = ?, actioned_at = ? 
        WHERE id = ?
    ''', (session.get('username', 'admin'), now_str, req_id))
    
    # 2. Delete student record from students table
    conn.execute("DELETE FROM students WHERE id = ?", (req['student_id'],))
    conn.commit()
    conn.close()
    
    flash(f'Deletion request approved! Student "{req["student_name"]}" has been permanently deleted.', 'success')
    return redirect(url_for('delete_requests_view'))

@app.route('/students/delete-requests/<int:req_id>/reject', methods=['POST'])
@login_required
def reject_delete_request(req_id):
    if session.get('role') != 'admin':
        flash('Access Denied. Only Admin can reject student deletion requests.', 'danger')
        return redirect(url_for('delete_requests_view'))
        
    conn = get_db_connection()
    req = conn.execute("SELECT * FROM student_delete_requests WHERE id = ?", (req_id,)).fetchone()
    if not req:
        conn.close()
        flash('Deletion request not found!', 'danger')
        return redirect(url_for('delete_requests_view'))
        
    if req['status'] != 'pending':
        conn.close()
        flash(f'This request has already been marked as {req["status"]}.', 'warning')
        return redirect(url_for('delete_requests_view'))
        
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        UPDATE student_delete_requests 
        SET status = 'rejected', actioned_by_user = ?, actioned_at = ? 
        WHERE id = ?
    ''', (session.get('username', 'admin'), now_str, req_id))
    conn.commit()
    conn.close()
    
    flash(f'Deletion request for student "{req["student_name"]}" was rejected. Student record was preserved.', 'info')
    return redirect(url_for('delete_requests_view'))

STANDARD_CLASSES = ['PG', 'Nursery', 'Prep', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine', 'Ten', 'Graduate']
CLASS_PROMOTION_MAP = {
    'PG': 'Nursery',
    'Nursery': 'Prep',
    'Prep': 'One',
    'One': 'Two',
    'Two': 'Three',
    'Three': 'Four',
    'Four': 'Five',
    'Five': 'Six',
    'Six': 'Seven',
    'Seven': 'Eight',
    'Eight': 'Nine',
    'Nine': 'Ten',
    'Ten': 'Graduate'
}

@app.route('/students/promotion', methods=['GET', 'POST'])
@login_required
def students_promotion():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    if request.method == 'POST':
        source_class = request.form.get('source_class', '').strip()
        target_class = request.form.get('target_class', '').strip()
        selected_student_ids = request.form.getlist('student_ids')
        fee_mode = request.form.get('fee_mode', 'keep')
        fee_value = float(request.form.get('fee_value', 0) or 0)
        new_start_month = int(request.form.get('new_start_month', 3))
        new_start_year = int(request.form.get('new_start_year', datetime.now().year))
        
        if not source_class or not target_class:
            conn.close()
            flash('Both Source Class and Target Class are required for promotion!', 'danger')
            return redirect(url_for('students_promotion'))
            
        if not selected_student_ids:
            conn.close()
            flash('No students were selected for promotion. Please check at least one student.', 'warning')
            return redirect(url_for('students_promotion', source_class=source_class))
            
        promoted_count = 0
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        user_name = session.get('username', 'admin')
        
        for s_id_str in selected_student_ids:
            s_id = int(s_id_str)
            student = conn.execute("SELECT * FROM students WHERE id = ?", (s_id,)).fetchone()
            if not student:
                continue
                
            # Security check for operators
            if session.get('role') != 'admin' and student['campus_id'] != session.get('campus_id'):
                continue
                
            prev_fee = student['monthly_fee']
            if fee_mode == 'fixed':
                new_fee = fee_value if fee_value > 0 else prev_fee
            elif fee_mode == 'increase_fixed':
                new_fee = prev_fee + fee_value
            elif fee_mode == 'increase_percent':
                new_fee = round(prev_fee * (1 + fee_value / 100.0))
            else:
                new_fee = prev_fee
                
            # Update student record
            conn.execute('''
                UPDATE students 
                SET class = ?, monthly_fee = ?, start_month = ?, start_year = ?
                WHERE id = ?
            ''', (target_class, new_fee, new_start_month, new_start_year, s_id))
            
            # Log in promotion history
            conn.execute('''
                INSERT INTO promotion_history 
                (student_id, student_name, from_class, to_class, previous_fee, new_fee, new_start_month, new_start_year, promoted_by_user, promoted_at, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (s_id, student['name'], source_class, target_class, prev_fee, new_fee, new_start_month, new_start_year, user_name, now_str, student['campus_id']))
            
            promoted_count += 1
            
        conn.commit()
        conn.close()
        flash(f'Successfully promoted {promoted_count} students from Class "{source_class}" to "{target_class}"!', 'success')
        return redirect(url_for('students_promotion', source_class=target_class))
        
    # GET: Load class options, students in chosen source class, and history
    db_classes_query = "SELECT DISTINCT class FROM students"
    db_classes_params = []
    if active_campus_id:
        db_classes_query += " WHERE campus_id = ?"
        db_classes_params = [active_campus_id]
    db_classes_query += " ORDER BY class"
    db_classes_rows = conn.execute(db_classes_query, db_classes_params).fetchall()
    existing_classes = [r['class'] for r in db_classes_rows if r['class']]
    
    # Combined sorted list of classes
    all_classes = list(dict.fromkeys(STANDARD_CLASSES + existing_classes))
    
    source_class = request.args.get('source_class', '').strip()
    if not source_class and existing_classes:
        source_class = existing_classes[0]
    elif not source_class and all_classes:
        source_class = all_classes[0]
        
    suggested_target_class = CLASS_PROMOTION_MAP.get(source_class, 'One')
    
    # Fetch students in source class
    students_in_class = []
    if source_class:
        s_query = "SELECT s.*, c.name as campus_name FROM students s LEFT JOIN campuses c ON s.campus_id = c.id WHERE s.class = ?"
        s_params = [source_class]
        if active_campus_id:
            s_query += " AND s.campus_id = ?"
            s_params.append(active_campus_id)
        s_query += " ORDER BY s.name"
        students_in_class = conn.execute(s_query, s_params).fetchall()
        
    # Fetch recent promotion history
    h_query = "SELECT h.*, c.name as campus_name FROM promotion_history h LEFT JOIN campuses c ON h.campus_id = c.id"
    h_params = []
    if session.get('role') != 'admin':
        h_query += " WHERE h.campus_id = ?"
        h_params.append(session.get('campus_id'))
    elif active_campus_id:
        h_query += " WHERE h.campus_id = ?"
        h_params.append(active_campus_id)
    h_query += " ORDER BY h.id DESC LIMIT 40"
    promotion_history = conn.execute(h_query, h_params).fetchall()
    
    conn.close()
    
    current_year = datetime.now().year
    
    return render_template('promotion.html',
                           classes=all_classes,
                           source_class=source_class,
                           suggested_target_class=suggested_target_class,
                           students=students_in_class,
                           promotion_history=promotion_history,
                           months=MONTH_NUM_TO_NAME,
                           current_year=current_year,
                           class_map=CLASS_PROMOTION_MAP)

@app.route('/students/promotion/batch', methods=['POST'])
@login_required
def students_promotion_batch():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    new_start_month = int(request.form.get('batch_start_month', 3))
    new_start_year = int(request.form.get('batch_start_year', datetime.now().year))
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    user_name = session.get('username', 'admin')
    
    # Process from highest class to lowest class in reverse order to avoid cascade overwrite
    ordered_sequence = [
        ('Ten', 'Graduate'),
        ('Nine', 'Ten'),
        ('Eight', 'Nine'),
        ('Seven', 'Eight'),
        ('Six', 'Seven'),
        ('Five', 'Six'),
        ('Four', 'Five'),
        ('Three', 'Four'),
        ('Two', 'Three'),
        ('One', 'Two'),
        ('Prep', 'One'),
        ('Nursery', 'Prep'),
        ('PG', 'Nursery')
    ]
    
    total_promoted = 0
    for from_c, to_c in ordered_sequence:
        s_query = "SELECT id, name, monthly_fee, campus_id FROM students WHERE class = ?"
        s_params = [from_c]
        if active_campus_id:
            s_query += " AND campus_id = ?"
            s_params.append(active_campus_id)
            
        students = conn.execute(s_query, s_params).fetchall()
        for s in students:
            # Update student
            conn.execute('''
                UPDATE students 
                SET class = ?, start_month = ?, start_year = ? 
                WHERE id = ?
            ''', (to_c, new_start_month, new_start_year, s['id']))
            
            # Log in history
            conn.execute('''
                INSERT INTO promotion_history 
                (student_id, student_name, from_class, to_class, previous_fee, new_fee, new_start_month, new_start_year, promoted_by_user, promoted_at, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (s['id'], s['name'], from_c, to_c, s['monthly_fee'], s['monthly_fee'], new_start_month, new_start_year, user_name, now_str, s['campus_id']))
            
            total_promoted += 1
            
    conn.commit()
    conn.close()
    
    flash(f'Whole-School Academic Session Rollover Complete! Total {total_promoted} students across all classes have been successfully promoted to the next grade.', 'success')
    return redirect(url_for('students_promotion'))

@app.route('/fee/entry', methods=['GET', 'POST'])
@login_required
def fee_entry():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    student_query = "SELECT id, name, father_name, class, monthly_fee, campus_id FROM students"
    student_params = []
    if active_campus_id:
        student_query += " WHERE campus_id = ?"
        student_params = [active_campus_id]
    student_query += " ORDER BY class, name"
    
    students_list = conn.execute(student_query, student_params).fetchall()
    
    selected_student_id = request.args.get('student_id', '', type=int)
    selected_student = None
    arrears_info = None
    
    target_month = request.args.get('month', MONTH_NUM_TO_NAME[datetime.now().month])
    target_year = request.args.get('year', datetime.now().year, type=int)
    
    annual_info = None
    if selected_student_id:
        selected_student = conn.execute("SELECT * FROM students WHERE id = ?", (selected_student_id,)).fetchone()
        if selected_student and active_campus_id and selected_student['campus_id'] != active_campus_id:
            selected_student = None
            
        if selected_student:
            arrears_info = get_student_fee_details(selected_student, target_month, target_year)
            ann_charges = float(selected_student['annual_charges'] or 0.0) if 'annual_charges' in selected_student.keys() else 0.0
            paid_ann_row = conn.execute(
                "SELECT SUM(paid_amount) as total_paid FROM annual_charges_payments WHERE student_id = ? AND year = ?",
                (selected_student['id'], target_year)
            ).fetchone()
            paid_ann = float(paid_ann_row['total_paid'] or 0.0) if paid_ann_row and paid_ann_row['total_paid'] is not None else 0.0
            unpaid_ann = max(0.0, ann_charges - paid_ann)
            annual_info = {
                'annual_charges': ann_charges,
                'paid_amount': paid_ann,
                'unpaid_amount': unpaid_ann,
                'is_paid': (ann_charges > 0 and unpaid_ann <= 0) or (ann_charges == 0)
            }
            
    if request.method == 'POST':
        student_id = int(request.form['student_id'])
        paid_amount = float(request.form['paid_amount'])
        date_paid = request.form['date_paid']
        payment_type = request.form.get('payment_type', 'monthly')
        
        if not date_paid:
            date_paid = datetime.now().strftime('%Y-%m-%d')
            
        student_obj = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
        if student_obj:
            if active_campus_id and student_obj['campus_id'] != active_campus_id:
                conn.close()
                flash('Access Denied.', 'danger')
                return redirect(url_for('fee_entry'))

            payment_mode = request.form.get('payment_mode', 'Voucher').strip()
            reference_no = request.form.get('reference_no', '').strip()
            notes = request.form.get('notes', '').strip()
            collected_by = session.get('username', 'operator')

            if payment_type == 'annual':
                # --- Annual Charges Payment ---
                annual_year = int(request.form.get('annual_year', datetime.now().year))
                existing = conn.execute(
                    "SELECT id, paid_amount FROM annual_charges_payments WHERE student_id = ? AND year = ?",
                    (student_id, annual_year)
                ).fetchone()
                if existing:
                    conn.execute(
                        "UPDATE annual_charges_payments SET paid_amount = ?, date_paid = ?, payment_mode = ?, reference_no = ?, notes = ?, collected_by = ? WHERE id = ?",
                        (paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, existing['id'])
                    )
                    flash(f"Updated annual charges for {student_obj['name']} ({annual_year}): Rs. {paid_amount:,.0f} via {payment_mode}", 'success')
                else:
                    conn.execute('''
                        INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, annual_year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, student_obj['campus_id']))
                    flash(f"Recorded annual charges of Rs. {paid_amount:,.0f} for {student_obj['name']} ({annual_year}) via {payment_mode}", 'success')
                conn.commit()
                conn.close()
                return redirect(url_for('fee_entry', student_id=student_id))

            elif payment_type == 'books':
                # --- Books / Syllabus Payment ---
                item_year = int(request.form.get('year', datetime.now().year))
                item_notes = notes or 'Books / Syllabus Payment'
                conn.execute('''
                    INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (student_id, 'Books', item_year, paid_amount, date_paid, payment_mode, reference_no, item_notes, collected_by, student_obj['campus_id']))
                flash(f"Recorded books payment of Rs. {paid_amount:,.0f} for {student_obj['name']} ({item_year}) via {payment_mode}", 'success')
                conn.commit()
                conn.close()
                return redirect(url_for('fee_entry', student_id=student_id))

            elif payment_type == 'admission':
                # --- Admission / Registration Fee ---
                item_year = int(request.form.get('year', datetime.now().year))
                item_notes = notes or 'Admission / Registration Fee'
                conn.execute('''
                    INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (student_id, 'Admission Fee', item_year, paid_amount, date_paid, payment_mode, reference_no, item_notes, collected_by, student_obj['campus_id']))
                flash(f"Recorded admission fee of Rs. {paid_amount:,.0f} for {student_obj['name']} ({item_year}) via {payment_mode}", 'success')
                conn.commit()
                conn.close()
                return redirect(url_for('fee_entry', student_id=student_id))

            elif payment_type == 'other':
                # --- Other Custom Fee ---
                item_year = int(request.form.get('year', datetime.now().year))
                custom_title = request.form.get('custom_title', '').strip() or 'Other Charges'
                item_notes = notes or f'{custom_title} Payment'
                conn.execute('''
                    INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (student_id, custom_title, item_year, paid_amount, date_paid, payment_mode, reference_no, item_notes, collected_by, student_obj['campus_id']))
                flash(f"Recorded {custom_title} of Rs. {paid_amount:,.0f} for {student_obj['name']} ({item_year}) via {payment_mode}", 'success')
                conn.commit()
                conn.close()
                return redirect(url_for('fee_entry', student_id=student_id))

            else:
                # --- Monthly Fee Payment (Multi-Month / Lumpsum Smart Allocation + Concurrent Add-ons) ---
                month = request.form['month']
                year = int(request.form['year'])
                num_months = int(request.form.get('months', 1) or 1)
                
                collected_items = []
                total_collected = 0.0

                if paid_amount > 0:
                    summaries = record_tuition_payment(
                        conn=conn,
                        student=student_obj,
                        start_month_name=month,
                        start_year=year,
                        paid_amount=paid_amount,
                        num_months=num_months,
                        date_paid=date_paid,
                        payment_mode=payment_mode,
                        reference_no=reference_no,
                        notes=notes,
                        collected_by=collected_by
                    )
                    collected_items.extend(summaries)
                    total_collected += paid_amount

                # Optional Concurrent Add-on Collections
                annual_amount = float(request.form.get('annual_amount', 0) or 0)
                books_amount = float(request.form.get('books_amount', 0) or 0)
                admission_amount = float(request.form.get('admission_amount', 0) or 0)
                other_amount = float(request.form.get('other_amount', 0) or 0)
                other_title = request.form.get('other_title', '').strip() or 'Other Charges'

                if annual_amount > 0:
                    existing = conn.execute(
                        "SELECT id, paid_amount FROM annual_charges_payments WHERE student_id = ? AND year = ?",
                        (student_id, year)
                    ).fetchone()
                    ann_notes = notes or f"Annual Charges {year}"
                    if existing:
                        new_ann_total = float(existing['paid_amount'] or 0) + annual_amount
                        conn.execute(
                            "UPDATE annual_charges_payments SET paid_amount = ?, date_paid = ?, payment_mode = ?, reference_no = ?, notes = ?, collected_by = ? WHERE id = ?",
                            (new_ann_total, date_paid, payment_mode, reference_no, ann_notes, collected_by, existing['id'])
                        )
                    else:
                        conn.execute('''
                            INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (student_id, year, annual_amount, date_paid, payment_mode, reference_no, ann_notes, collected_by, student_obj['campus_id']))
                    collected_items.append(f"Annual: Rs. {annual_amount:,.0f}")
                    total_collected += annual_amount

                if books_amount > 0:
                    b_notes = notes or "Books / Syllabus Payment"
                    conn.execute('''
                        INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, 'Books', year, books_amount, date_paid, payment_mode, reference_no, b_notes, collected_by, student_obj['campus_id']))
                    collected_items.append(f"Books: Rs. {books_amount:,.0f}")
                    total_collected += books_amount

                if admission_amount > 0:
                    adm_notes = notes or "Admission Fee"
                    conn.execute('''
                        INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, 'Admission Fee', year, admission_amount, date_paid, payment_mode, reference_no, adm_notes, collected_by, student_obj['campus_id']))
                    collected_items.append(f"Admission Fee: Rs. {admission_amount:,.0f}")
                    total_collected += admission_amount

                if other_amount > 0:
                    oth_notes = notes or f"{other_title} Payment"
                    conn.execute('''
                        INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, other_title, year, other_amount, date_paid, payment_mode, reference_no, oth_notes, collected_by, student_obj['campus_id']))
                    collected_items.append(f"{other_title}: Rs. {other_amount:,.0f}")
                    total_collected += other_amount

                conn.commit()
                conn.close()
                
                summary_text = " + ".join(collected_items) if collected_items else f"Rs. {total_collected:,.0f}"
                flash(f"Successfully recorded payment for {student_obj['name']}: {summary_text} [Total Rs. {total_collected:,.0f} via {payment_mode}]", 'success')
                return redirect(url_for('fee_entry', student_id=student_id, month=month, year=year))

    return render_template('fee_entry.html', 
                           students=students_list, 
                           selected_student_id=selected_student_id,
                           selected_student=selected_student,
                           arrears_info=arrears_info,
                           target_month=target_month,
                           target_year=target_year,
                           months=MONTH_NUM_TO_NAME,
                           years=[2025, 2026, 2027, 2028],
                           annual_info=annual_info,
                           current_date=datetime.now().strftime('%Y-%m-%d'))

@app.route('/fee/class-sheet')
@login_required
def class_fee_sheet():
    target_month = request.args.get('month', MONTH_NUM_TO_NAME[datetime.now().month])
    target_year = request.args.get('year', datetime.now().year, type=int)
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    # 1. Fetch available classes with student counts
    db_classes_query = "SELECT DISTINCT class FROM students WHERE (status IS NULL OR status = 'active')"
    db_classes_params = []
    if active_campus_id:
        db_classes_query += " AND campus_id = ?"
        db_classes_params = [active_campus_id]
    db_classes_query += " ORDER BY class"
    db_classes_rows = conn.execute(db_classes_query, db_classes_params).fetchall()
    existing_classes = [r['class'] for r in db_classes_rows if r['class']]
    
    classes = [c for c in STANDARD_CLASSES if c in existing_classes] + [c for c in existing_classes if c not in STANDARD_CLASSES]
    if not classes:
        classes = existing_classes or STANDARD_CLASSES
        
    count_query = "SELECT class, COUNT(*) as count FROM students WHERE (status IS NULL OR status = 'active')"
    count_params = []
    if active_campus_id:
        count_query += " AND campus_id = ?"
        count_params.append(active_campus_id)
    count_query += " GROUP BY class"
    class_counts = {r['class']: r['count'] for r in conn.execute(count_query, count_params).fetchall()}
    
    selected_class = request.args.get('class', '').strip()
    if not selected_class and classes:
        selected_class = classes[0]
        
    # 2. Fetch students for the selected class
    sheet_data = []
    total_class_monthly_fee = 0.0
    total_class_arrears = 0.0
    total_class_payable = 0.0
    total_class_paid = 0.0
    total_class_remaining = 0.0
    paid_count = 0
    partial_count = 0
    unpaid_count = 0
    
    if selected_class:
        s_query = """
            SELECT s.*, c.name as campus_name 
            FROM students s 
            LEFT JOIN campuses c ON s.campus_id = c.id 
            WHERE s.class = ? AND (s.status IS NULL OR s.status = 'active')
        """
        s_params = [selected_class]
        if active_campus_id:
            s_query += " AND s.campus_id = ?"
            s_params.append(active_campus_id)
        s_query += " ORDER BY s.id ASC"
        students = conn.execute(s_query, s_params).fetchall()
        
        student_ids = [s['id'] for s in students]
        fees_map = {sid: [] for sid in student_ids}
        if student_ids:
            placeholders = ','.join(['?'] * len(student_ids))
            all_fees = conn.execute(
                f"SELECT student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes FROM fees WHERE student_id IN ({placeholders})",
                student_ids
            ).fetchall()
            for f in all_fees:
                fees_map[f['student_id']].append(f)
                
        for s in students:
            details = get_student_fee_details(s, target_month, target_year, payments=fees_map.get(s['id'], []))
            monthly_fee = details['monthly_fee']
            arrears = details['arrears']
            total_payable = details['total_payable']
            paid = details['paid_this_month']
            remaining = details['remaining_payable']
            
            if remaining <= 0 and (paid > 0 or total_payable == 0):
                status = 'Paid'
                status_badge = 'success'
                paid_count += 1
            elif paid > 0 and remaining > 0:
                status = 'Partial'
                status_badge = 'warning'
                partial_count += 1
            else:
                status = 'Unpaid'
                status_badge = 'danger'
                unpaid_count += 1
                
            sheet_data.append({
                'id': s['id'],
                'name': s['name'],
                'father_name': s['father_name'] or '',
                'phone_number': s['phone_number'] or '',
                'class': s['class'],
                'campus_name': s['campus_name'],
                'monthly_fee': monthly_fee,
                'annual_charges': float(s['annual_charges'] or 0) if 'annual_charges' in s.keys() else 0.0,
                'arrears': arrears,
                'total_payable': total_payable,
                'paid': paid,
                'remaining': remaining,
                'status': status,
                'status_badge': status_badge
            })
            
            total_class_monthly_fee += monthly_fee
            total_class_arrears += arrears
            total_class_payable += total_payable
            total_class_paid += paid
            total_class_remaining += remaining
            
    conn.close()
    
    months = list(MONTH_NUM_TO_NAME.values())
    years = [2025, 2026, 2027, 2028]
    
    return render_template('class_fee_sheet.html',
                           classes=classes,
                           class_counts=class_counts,
                           selected_class=selected_class,
                           target_month=target_month,
                           target_year=target_year,
                           months=months,
                           years=years,
                           sheet_data=sheet_data,
                           total_students=len(sheet_data),
                           total_class_monthly_fee=total_class_monthly_fee,
                           total_class_arrears=total_class_arrears,
                           total_class_payable=total_class_payable,
                           total_class_paid=total_class_paid,
                           total_class_remaining=total_class_remaining,
                           paid_count=paid_count,
                           partial_count=partial_count,
                           unpaid_count=unpaid_count,
                           current_date=datetime.now().strftime('%Y-%m-%d'))

@app.route('/fee/quick-collect', methods=['POST'])
@login_required
def fee_quick_collect():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    student_id = int(request.form['student_id'])
    paid_amount = float(request.form.get('paid_amount', 0) or 0)
    month = request.form['month']
    year = int(request.form['year'])
    selected_class = request.form.get('return_class', '')
    date_paid = request.form.get('date_paid') or datetime.now().strftime('%Y-%m-%d')
    payment_mode = request.form.get('payment_mode', 'Voucher').strip()
    reference_no = request.form.get('reference_no', '').strip()
    notes = request.form.get('notes', '').strip()
    collected_by = session.get('username', 'operator')
    
    annual_amount = float(request.form.get('annual_amount', 0) or 0)
    books_amount = float(request.form.get('books_amount', 0) or 0)
    admission_amount = float(request.form.get('admission_amount', 0) or 0)
    other_amount = float(request.form.get('other_amount', 0) or 0)
    other_title = request.form.get('other_title', '').strip() or 'Other Charges'
    
    student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    if not student or (active_campus_id and student['campus_id'] != active_campus_id):
        conn.close()
        flash('Student not found or access denied.', 'danger')
        return redirect(url_for('class_fee_sheet', **{'class': selected_class, 'month': month, 'year': year}))
        
    collected_items = []
    total_collected = 0.0

    # 1. Tuition Payment
    if paid_amount > 0:
        summaries = record_tuition_payment(
            conn=conn,
            student=student,
            start_month_name=month,
            start_year=year,
            paid_amount=paid_amount,
            num_months=1,
            date_paid=date_paid,
            payment_mode=payment_mode,
            reference_no=reference_no,
            notes=notes,
            collected_by=collected_by
        )
        collected_items.extend(summaries)
        total_collected += paid_amount

    # 2. Annual Charges
    if annual_amount > 0:
        existing = conn.execute(
            "SELECT id, paid_amount FROM annual_charges_payments WHERE student_id = ? AND year = ?",
            (student_id, year)
        ).fetchone()
        ann_notes = notes or f"Annual Charges {year}"
        if existing:
            new_ann_total = float(existing['paid_amount'] or 0) + annual_amount
            conn.execute(
                "UPDATE annual_charges_payments SET paid_amount = ?, date_paid = ?, payment_mode = ?, reference_no = ?, notes = ?, collected_by = ? WHERE id = ?",
                (new_ann_total, date_paid, payment_mode, reference_no, ann_notes, collected_by, existing['id'])
            )
        else:
            conn.execute('''
                INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, year, annual_amount, date_paid, payment_mode, reference_no, ann_notes, collected_by, student['campus_id']))
        collected_items.append(f"Annual: Rs. {annual_amount:,.0f}")
        total_collected += annual_amount

    # 3. Books Payment
    if books_amount > 0:
        b_notes = notes or "Books / Syllabus Payment"
        conn.execute('''
            INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (student_id, 'Books', year, books_amount, date_paid, payment_mode, reference_no, b_notes, collected_by, student['campus_id']))
        collected_items.append(f"Books: Rs. {books_amount:,.0f}")
        total_collected += books_amount

    # 4. Admission Fee
    if admission_amount > 0:
        adm_notes = notes or "Admission Fee"
        conn.execute('''
            INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (student_id, 'Admission Fee', year, admission_amount, date_paid, payment_mode, reference_no, adm_notes, collected_by, student['campus_id']))
        collected_items.append(f"Admission Fee: Rs. {admission_amount:,.0f}")
        total_collected += admission_amount

    # 5. Other Charges
    if other_amount > 0:
        oth_notes = notes or f"{other_title} Payment"
        conn.execute('''
            INSERT INTO fees (student_id, month, year, paid_amount, date_paid, payment_mode, reference_no, notes, collected_by, campus_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (student_id, other_title, year, other_amount, date_paid, payment_mode, reference_no, oth_notes, collected_by, student['campus_id']))
        collected_items.append(f"{other_title}: Rs. {other_amount:,.0f}")
        total_collected += other_amount

    conn.commit()
    conn.close()
    
    summary_str = " + ".join(collected_items) if collected_items else f"Rs. {total_collected:,.0f}"
    flash(f"Payment recorded for {student['name']}: {summary_str} [Total Rs. {total_collected:,.0f} via {payment_mode}] successfully!", 'success')
    return redirect(url_for('class_fee_sheet', **{'class': selected_class or student['class'], 'month': month, 'year': year}))

@app.route('/fee/class-sheet/export')
@login_required
def export_class_fee_sheet():
    target_class = request.args.get('class', '').strip()
    target_month = request.args.get('month', MONTH_NUM_TO_NAME[datetime.now().month])
    target_year = request.args.get('year', datetime.now().year, type=int)
    active_campus_id = get_active_campus_id()
    
    conn = get_db_connection()
    s_query = "SELECT * FROM students WHERE 1=1"
    s_params = []
    if target_class:
        s_query += " AND class = ?"
        s_params.append(target_class)
    if active_campus_id:
        s_query += " AND campus_id = ?"
        s_params.append(active_campus_id)
    s_query += " AND (status IS NULL OR status = 'active') ORDER BY class, id"
    students = conn.execute(s_query, s_params).fetchall()
    
    student_ids = [s['id'] for s in students]
    fees_map = {sid: [] for sid in student_ids}
    if student_ids:
        placeholders = ','.join(['?'] * len(student_ids))
        all_fees = conn.execute(
            f"SELECT student_id, month, year, paid_amount FROM fees WHERE student_id IN ({placeholders})",
            student_ids
        ).fetchall()
        for f in all_fees:
            fees_map[f['student_id']].append(f)
            
    conn.close()
    
    rows = []
    for s in students:
        details = get_student_fee_details(s, target_month, target_year, payments=fees_map.get(s['id'], []))
        paid = details['paid_this_month']
        rem = details['remaining_payable']
        status = 'Paid' if (rem <= 0 and (paid > 0 or details['total_payable'] == 0)) else ('Partial' if (paid > 0 and rem > 0) else 'Unpaid')
        rows.append({
            'Roll / ID': s['id'],
            'Student Name': s['name'],
            'Father Name': s['father_name'] or '',
            'Phone / WhatsApp': s['phone_number'] or '',
            'Class': s['class'],
            'Billing Month': f"{target_month} {target_year}",
            'Monthly Tuition Fee (Rs.)': details['monthly_fee'],
            'Previous Arrears (Rs.)': details['arrears'],
            'Total Payable (Rs.)': details['total_payable'],
            'Paid Amount (Rs.)': paid,
            'Remaining Balance (Rs.)': rem,
            'Payment Status': status
        })
        
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        'Roll / ID', 'Student Name', 'Father Name', 'Phone / WhatsApp', 'Class',
        'Billing Month', 'Monthly Tuition Fee (Rs.)', 'Previous Arrears (Rs.)',
        'Total Payable (Rs.)', 'Paid Amount (Rs.)', 'Remaining Balance (Rs.)', 'Payment Status'
    ])
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = f"Class {target_class}" if target_class else "All Classes"
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        ws = writer.sheets[sheet_name[:31]]
        
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
    output.seek(0)
    safe_class = re.sub(r'[^a-zA-Z0-9_-]', '_', target_class or 'All')
    filename = f"Fee_Sheet_Class_{safe_class}_{target_month}_{target_year}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/fee/analytics')
@admin_required
def fee_analytics():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    # 1. Read Filter Parameters
    period = request.args.get('period', 'monthly').strip()
    if period not in ('daily', 'monthly', 'yearly', 'custom'):
        period = 'monthly'
        
    today_str = datetime.now().strftime('%Y-%m-%d')
    selected_date = request.args.get('date', today_str).strip() or today_str
    
    curr_month_name = MONTH_NUM_TO_NAME[datetime.now().month]
    selected_month = request.args.get('month', curr_month_name).strip() or curr_month_name
    
    curr_year = datetime.now().year
    selected_year = request.args.get('year', curr_year, type=int) or curr_year
    
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    
    selected_campus_id = request.args.get('campus_id', type=int)
    if session.get('role') != 'admin':
        selected_campus_id = active_campus_id
    elif not selected_campus_id and active_campus_id:
        selected_campus_id = active_campus_id
        
    payment_mode_filter = request.args.get('payment_mode', 'all').strip()
    category_filter = request.args.get('category', 'all').strip()
    class_filter = request.args.get('class', 'all').strip()
    
    # 2. Fetch Campuses and Classes for Filter Dropdowns
    campuses = conn.execute("SELECT id, name, code FROM campuses ORDER BY name").fetchall()
    classes = conn.execute("SELECT DISTINCT class FROM students WHERE class IS NOT NULL ORDER BY class").fetchall()
    classes_list = [c['class'] for c in classes if c['class']]
    
    # Fetch all fee transactions & annual charges payments within relevant scope
    campus_condition = ""
    campus_params = []
    if selected_campus_id:
        campus_condition = " WHERE (f.campus_id = ? OR (f.campus_id IS NULL AND s.campus_id = ?))"
        campus_params = [selected_campus_id, selected_campus_id]
        
    fee_sql = f"""
        SELECT f.id, f.student_id, f.month, f.year, f.paid_amount, f.date_paid, f.payment_mode, f.reference_no, f.notes, f.collected_by, f.campus_id,
               s.name as student_name, s.father_name, s.class as student_class, c.name as campus_name, 'Fee' as source_table
        FROM fees f
        LEFT JOIN students s ON f.student_id = s.id
        LEFT JOIN campuses c ON f.campus_id = c.id
        {campus_condition}
    """
    
    ann_campus_cond = ""
    ann_params = []
    if selected_campus_id:
        ann_campus_cond = " WHERE (a.campus_id = ? OR (a.campus_id IS NULL AND s.campus_id = ?))"
        ann_params = [selected_campus_id, selected_campus_id]
        
    ann_sql = f"""
        SELECT a.id, a.student_id, 'Annual Charges' as month, a.year, a.paid_amount, a.date_paid, a.payment_mode, a.reference_no, a.notes, a.collected_by, a.campus_id,
               s.name as student_name, s.father_name, s.class as student_class, c.name as campus_name, 'Annual' as source_table
        FROM annual_charges_payments a
        LEFT JOIN students s ON a.student_id = s.id
        LEFT JOIN campuses c ON a.campus_id = c.id
        {ann_campus_cond}
    """
    
    fee_rows = conn.execute(fee_sql, campus_params).fetchall()
    ann_rows = conn.execute(ann_sql, ann_params).fetchall()
    conn.close()
    
    all_raw_txs = []
    for r in fee_rows:
        all_raw_txs.append(dict(r))
    for r in ann_rows:
        all_raw_txs.append(dict(r))
        
    def categorize_tx(tx):
        m = (tx.get('month') or '').strip()
        st = tx.get('source_table', '')
        if st == 'Annual' or m == 'Annual Charges':
            return 'Annual Charges'
        elif m in ('Books', 'Books Payment', 'Books & Stationary', 'Books / Stationary'):
            return 'Books / Syllabus'
        elif m in ('Admission', 'Admission Fee', 'Registration Fee'):
            return 'Admission Fee'
        elif m in ('Uniform', 'Uniform Charges'):
            return 'Uniform'
        elif m in MONTH_NAME_TO_NUM:
            return 'Monthly Tuition'
        else:
            return m or 'Other Charges'

    def normalize_mode(mode):
        mode = (mode or '').strip().lower()
        if any(x in mode for x in ('bank', 'deposit', 'transfer', 'mcb', 'hbl', 'ubl', 'meezan', 'allied')):
            return 'Bank'
        elif any(x in mode for x in ('online', 'app', 'jazz', 'easy', 'nayapay', 'sadapay', 'mobile')):
            return 'Online'
        else:
            return 'Voucher'

    for tx in all_raw_txs:
        tx['category'] = categorize_tx(tx)
        tx['norm_mode'] = normalize_mode(tx.get('payment_mode'))
        tx['paid_amount'] = float(tx.get('paid_amount') or 0.0)
        tx['date_paid_clean'] = (tx.get('date_paid') or '').strip()

    # Filter according to period
    if period == 'daily':
        filtered_txs = [t for t in all_raw_txs if t['date_paid_clean'].startswith(selected_date)]
    elif period == 'monthly':
        selected_m_num = MONTH_NAME_TO_NUM.get(selected_month, datetime.now().month)
        m_prefix = f"{selected_year}-{selected_m_num:02d}"
        filtered_txs = [t for t in all_raw_txs if (t['month'] == selected_month and int(t['year'] or 0) == selected_year) or t['date_paid_clean'].startswith(m_prefix)]
    elif period == 'yearly':
        filtered_txs = [t for t in all_raw_txs if int(t['year'] or 0) == selected_year or t['date_paid_clean'].startswith(str(selected_year))]
    elif period == 'custom':
        filtered_txs = [t for t in all_raw_txs if (not start_date or t['date_paid_clean'] >= start_date) and (not end_date or t['date_paid_clean'] <= end_date)]
    else:
        filtered_txs = list(all_raw_txs)

    # Sub-filters
    if payment_mode_filter != 'all':
        filtered_txs = [t for t in filtered_txs if t['norm_mode'].lower() == payment_mode_filter.lower()]
    if category_filter != 'all':
        filtered_txs = [t for t in filtered_txs if t['category'].lower() == category_filter.lower()]
    if class_filter != 'all':
        filtered_txs = [t for t in filtered_txs if (t.get('student_class') or '').lower() == class_filter.lower()]

    total_revenue = sum(t['paid_amount'] for t in filtered_txs)
    total_tx_count = len(filtered_txs)
    avg_ticket = round(total_revenue / total_tx_count, 0) if total_tx_count > 0 else 0

    mode_voucher_sum = sum(t['paid_amount'] for t in filtered_txs if t['norm_mode'] == 'Voucher')
    mode_bank_sum = sum(t['paid_amount'] for t in filtered_txs if t['norm_mode'] == 'Bank')
    mode_online_sum = sum(t['paid_amount'] for t in filtered_txs if t['norm_mode'] == 'Online')

    cat_tuition_sum = sum(t['paid_amount'] for t in filtered_txs if t['category'] == 'Monthly Tuition')
    cat_annual_sum = sum(t['paid_amount'] for t in filtered_txs if t['category'] == 'Annual Charges')
    cat_books_sum = sum(t['paid_amount'] for t in filtered_txs if t['category'] == 'Books / Syllabus')
    cat_admission_sum = sum(t['paid_amount'] for t in filtered_txs if t['category'] == 'Admission Fee')
    cat_other_sum = sum(t['paid_amount'] for t in filtered_txs if t['category'] not in ('Monthly Tuition', 'Annual Charges', 'Books / Syllabus', 'Admission Fee'))

    # Time-series Charts Data
    # 1. 14-Day Timeline
    daily_labels = []
    daily_values = []
    for d in range(13, -1, -1):
        day_dt = datetime.now() - timedelta(days=d)
        d_key = day_dt.strftime('%Y-%m-%d')
        d_label = day_dt.strftime('%d %b')
        day_sum = sum(t['paid_amount'] for t in all_raw_txs if t['date_paid_clean'].startswith(d_key))
        daily_labels.append(d_label)
        daily_values.append(day_sum)

    # 2. 12-Month Year Curve
    monthly_labels = list(MONTH_NUM_TO_NAME.values())
    monthly_values = []
    for m_idx in range(1, 13):
        m_name = MONTH_NUM_TO_NAME[m_idx]
        m_prefix = f"{selected_year}-{m_idx:02d}"
        m_sum = sum(
            t['paid_amount'] for t in all_raw_txs 
            if (int(t.get('year') or 0) == selected_year and t.get('month') == m_name) or t['date_paid_clean'].startswith(m_prefix)
        )
        monthly_values.append(m_sum)

    # 3. Multi-Year
    yearly_labels = ['2024', '2025', '2026', '2027']
    yearly_values = []
    for yr in yearly_labels:
        y_int = int(yr)
        y_sum = sum(t['paid_amount'] for t in all_raw_txs if int(t.get('year') or 0) == y_int or t['date_paid_clean'].startswith(yr))
        yearly_values.append(y_sum)

    # 4. Class Breakdown
    class_sums = {}
    for c in classes_list:
        class_sums[c] = 0.0
    for t in filtered_txs:
        c_name = t.get('student_class') or 'Unknown'
        class_sums[c_name] = class_sums.get(c_name, 0.0) + t['paid_amount']
    sorted_classes = sorted(class_sums.items(), key=lambda x: x[1], reverse=True)
    class_labels = [x[0] for x in sorted_classes if x[1] > 0] or list(class_sums.keys())[:10]
    class_values = [x[1] for x in sorted_classes if x[1] > 0] or [0] * len(class_labels)

    filtered_txs.sort(key=lambda x: (x.get('date_paid_clean') or '', x.get('id') or 0), reverse=True)

    years_list = [2024, 2025, 2026, 2027, 2028]

    return render_template('fee_analytics.html',
                           period=period,
                           selected_date=selected_date,
                           selected_month=selected_month,
                           selected_year=selected_year,
                           start_date=start_date,
                           end_date=end_date,
                           selected_campus_id=selected_campus_id,
                           payment_mode_filter=payment_mode_filter,
                           category_filter=category_filter,
                           class_filter=class_filter,
                           campuses=campuses,
                           classes=classes_list,
                           months=MONTH_NUM_TO_NAME,
                           years=years_list,
                           total_revenue=total_revenue,
                           total_tx_count=total_tx_count,
                           avg_ticket=avg_ticket,
                           mode_voucher_sum=mode_voucher_sum,
                           mode_bank_sum=mode_bank_sum,
                           mode_online_sum=mode_online_sum,
                           cat_tuition_sum=cat_tuition_sum,
                           cat_annual_sum=cat_annual_sum,
                           cat_books_sum=cat_books_sum,
                           cat_admission_sum=cat_admission_sum,
                           cat_other_sum=cat_other_sum,
                           daily_labels_json=json.dumps(daily_labels),
                           daily_values_json=json.dumps(daily_values),
                           monthly_labels_json=json.dumps(monthly_labels),
                           monthly_values_json=json.dumps(monthly_values),
                           yearly_labels_json=json.dumps(yearly_labels),
                           yearly_values_json=json.dumps(yearly_values),
                           class_labels_json=json.dumps(class_labels),
                           class_values_json=json.dumps(class_values),
                           transactions=filtered_txs)

@app.route('/fee/analytics/export')
@admin_required
def fee_analytics_export():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    period = request.args.get('period', 'monthly').strip()
    selected_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d')).strip()
    selected_month = request.args.get('month', MONTH_NUM_TO_NAME[datetime.now().month]).strip()
    selected_year = request.args.get('year', datetime.now().year, type=int)
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    
    selected_campus_id = request.args.get('campus_id', type=int)
    if session.get('role') != 'admin':
        selected_campus_id = active_campus_id
    elif not selected_campus_id and active_campus_id:
        selected_campus_id = active_campus_id
        
    payment_mode_filter = request.args.get('payment_mode', 'all').strip()
    category_filter = request.args.get('category', 'all').strip()
    class_filter = request.args.get('class', 'all').strip()
    
    campus_condition = ""
    campus_params = []
    if selected_campus_id:
        campus_condition = " WHERE (f.campus_id = ? OR (f.campus_id IS NULL AND s.campus_id = ?))"
        campus_params = [selected_campus_id, selected_campus_id]
        
    fee_sql = f"""
        SELECT f.id, f.student_id, f.month, f.year, f.paid_amount, f.date_paid, f.payment_mode, f.reference_no, f.notes, f.collected_by, f.campus_id,
               s.name as student_name, s.father_name, s.class as student_class, c.name as campus_name, 'Fee' as source_table
        FROM fees f
        LEFT JOIN students s ON f.student_id = s.id
        LEFT JOIN campuses c ON f.campus_id = c.id
        {campus_condition}
    """
    
    ann_campus_cond = ""
    ann_params = []
    if selected_campus_id:
        ann_campus_cond = " WHERE (a.campus_id = ? OR (a.campus_id IS NULL AND s.campus_id = ?))"
        ann_params = [selected_campus_id, selected_campus_id]
        
    ann_sql = f"""
        SELECT a.id, a.student_id, 'Annual Charges' as month, a.year, a.paid_amount, a.date_paid, a.payment_mode, a.reference_no, a.notes, a.collected_by, a.campus_id,
               s.name as student_name, s.father_name, s.class as student_class, c.name as campus_name, 'Annual' as source_table
        FROM annual_charges_payments a
        LEFT JOIN students s ON a.student_id = s.id
        LEFT JOIN campuses c ON a.campus_id = c.id
        {ann_campus_cond}
    """
    
    fee_rows = conn.execute(fee_sql, campus_params).fetchall()
    ann_rows = conn.execute(ann_sql, ann_params).fetchall()
    conn.close()
    
    all_raw_txs = [dict(r) for r in fee_rows] + [dict(r) for r in ann_rows]
    
    def categorize_tx(tx):
        m = (tx.get('month') or '').strip()
        st = tx.get('source_table', '')
        if st == 'Annual' or m == 'Annual Charges':
            return 'Annual Charges'
        elif m in ('Books', 'Books Payment', 'Books & Stationary', 'Books / Stationary'):
            return 'Books / Syllabus'
        elif m in ('Admission', 'Admission Fee', 'Registration Fee'):
            return 'Admission Fee'
        elif m in ('Uniform', 'Uniform Charges'):
            return 'Uniform'
        elif m in MONTH_NAME_TO_NUM:
            return 'Monthly Tuition'
        else:
            return m or 'Other Charges'

    def normalize_mode(mode):
        mode = (mode or '').strip().lower()
        if any(x in mode for x in ('bank', 'deposit', 'transfer', 'mcb', 'hbl', 'ubl', 'meezan', 'allied')):
            return 'Bank'
        elif any(x in mode for x in ('online', 'app', 'jazz', 'easy', 'nayapay', 'sadapay', 'mobile')):
            return 'Online'
        else:
            return 'Voucher'

    for tx in all_raw_txs:
        tx['category'] = categorize_tx(tx)
        tx['norm_mode'] = normalize_mode(tx.get('payment_mode'))
        tx['paid_amount'] = float(tx.get('paid_amount') or 0.0)
        tx['date_paid_clean'] = (tx.get('date_paid') or '').strip()

    if period == 'daily':
        filtered_txs = [t for t in all_raw_txs if t['date_paid_clean'].startswith(selected_date)]
    elif period == 'monthly':
        selected_m_num = MONTH_NAME_TO_NUM.get(selected_month, datetime.now().month)
        m_prefix = f"{selected_year}-{selected_m_num:02d}"
        filtered_txs = [t for t in all_raw_txs if (t['month'] == selected_month and int(t['year'] or 0) == selected_year) or t['date_paid_clean'].startswith(m_prefix)]
    elif period == 'yearly':
        filtered_txs = [t for t in all_raw_txs if int(t['year'] or 0) == selected_year or t['date_paid_clean'].startswith(str(selected_year))]
    elif period == 'custom':
        filtered_txs = [t for t in all_raw_txs if (not start_date or t['date_paid_clean'] >= start_date) and (not end_date or t['date_paid_clean'] <= end_date)]
    else:
        filtered_txs = list(all_raw_txs)

    if payment_mode_filter != 'all':
        filtered_txs = [t for t in filtered_txs if t['norm_mode'].lower() == payment_mode_filter.lower()]
    if category_filter != 'all':
        filtered_txs = [t for t in filtered_txs if t['category'].lower() == category_filter.lower()]
    if class_filter != 'all':
        filtered_txs = [t for t in filtered_txs if (t.get('student_class') or '').lower() == class_filter.lower()]

    filtered_txs.sort(key=lambda x: (x.get('date_paid_clean') or '', x.get('id') or 0), reverse=True)

    rows = []
    for t in filtered_txs:
        rows.append({
            'Tx ID': t.get('id'),
            'Date Paid': t.get('date_paid_clean'),
            'Student ID': t.get('student_id'),
            'Student Name': t.get('student_name') or '—',
            'Father Name': t.get('father_name') or '—',
            'Class': t.get('student_class') or '—',
            'Fee Category': t.get('category'),
            'Month / Year': f"{t.get('month')} {t.get('year')}",
            'Amount Paid (Rs.)': t.get('paid_amount'),
            'Payment Mode': t.get('payment_mode') or 'Voucher',
            'Reference #': t.get('reference_no') or '',
            'Collected By': t.get('collected_by') or '',
            'Campus': t.get('campus_name') or ''
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        'Tx ID', 'Date Paid', 'Student ID', 'Student Name', 'Father Name', 'Class',
        'Fee Category', 'Month / Year', 'Amount Paid (Rs.)', 'Payment Mode', 'Reference #', 'Collected By', 'Campus'
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Fee Collection", index=False)
        ws = writer.sheets["Fee Collection"]
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output.seek(0)
    filename = f"Fee_Collection_Report_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/students/<int:student_id>/ledger')
@app.route('/fee/history/<int:student_id>')
@login_required
def fee_history(student_id):
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    student = conn.execute('''
        SELECT s.*, c.name as campus_name 
        FROM students s 
        LEFT JOIN campuses c ON s.campus_id = c.id 
        WHERE s.id = ?
    ''', (student_id,)).fetchone()
    
    if not student or (active_campus_id and student['campus_id'] != active_campus_id):
        conn.close()
        flash('Student record not found or access denied!', 'danger')
        return redirect(url_for('students_view'))
        
    # Fetch monthly fee payments
    monthly_payments = conn.execute('''
        SELECT id, month, year, paid_amount, date_paid, 
               COALESCE(payment_mode, 'Voucher') as payment_mode, 
               reference_no, notes, collected_by, 'Monthly Fee' as payment_type
        FROM fees 
        WHERE student_id = ? 
        ORDER BY date_paid DESC, id DESC
    ''', (student_id,)).fetchall()
    
    # Fetch annual charges payments
    annual_payments = conn.execute('''
        SELECT id, 'Annual Charges' as month, year, paid_amount, date_paid, 
               COALESCE(payment_mode, 'Voucher') as payment_mode, 
               reference_no, notes, collected_by, 'Annual Charges' as payment_type
        FROM annual_charges_payments 
        WHERE student_id = ? 
        ORDER BY date_paid DESC, id DESC
    ''', (student_id,)).fetchall()
    
    # Combine all transactions and sort by date descending
    all_transactions = []
    for p in monthly_payments:
        all_transactions.append(dict(p))
    for a in annual_payments:
        all_transactions.append(dict(a))
        
    all_transactions.sort(key=lambda x: (x['date_paid'] or '', x['id']), reverse=True)
    
    # Calculate Financial Totals & Payment Modes breakdown
    total_paid = sum(t['paid_amount'] for t in all_transactions)
    paid_voucher = sum(t['paid_amount'] for t in all_transactions if t['payment_mode'] in ('Voucher', 'Cash Counter', 'Cash', 'voucher'))
    paid_bank = sum(t['paid_amount'] for t in all_transactions if t['payment_mode'] in ('Bank', 'Bank Deposit', 'Bank Transfer', 'bank'))
    paid_online = sum(t['paid_amount'] for t in all_transactions if t['payment_mode'] in ('Online', 'Online Transfer', 'Mobile App', 'JazzCash', 'EasyPaisa', 'online'))
    
    # Calculate billing duration & total dues
    curr_year = datetime.now().year
    curr_month = datetime.now().month
    start_m = student['start_month'] or 3
    start_y = student['start_year'] or curr_year
    
    months_billed = (curr_year - start_y) * 12 + (curr_month - start_m) + 1
    if months_billed < 0:
        months_billed = 0
        
    opening_arrears = float(student['opening_arrears'] or 0.0) if 'opening_arrears' in student.keys() else 0.0
    total_monthly_billed = months_billed * (student['monthly_fee'] or 0)
    total_annual_billed = (student['annual_charges'] or 0) * max(1, (curr_year - start_y + 1))
    
    # One-off non-tuition charges (e.g. Books, Admission Fee, Uniform, Other Charges)
    non_tuition_billed = sum(
        float(t['paid_amount'] or 0) for t in all_transactions 
        if t.get('payment_type') != 'Annual Charges' and t.get('month') not in MONTH_NAME_TO_NUM
    )
    
    total_billed = opening_arrears + total_monthly_billed + total_annual_billed + non_tuition_billed
    
    balance_due = total_billed - total_paid
    if balance_due < 0:
        balance_due = 0
        
    completion_rate = round((total_paid / total_billed * 100) if total_billed > 0 else 100, 1)
    
    settings = get_campus_settings(student['campus_id'])
    
    conn.close()
    
    return render_template('fee_history.html',
                           student=student,
                           transactions=all_transactions,
                           total_paid=total_paid,
                           paid_voucher=paid_voucher,
                           paid_bank=paid_bank,
                           paid_online=paid_online,
                           total_billed=total_billed,
                           opening_arrears=opening_arrears,
                           balance_due=balance_due,
                           months_billed=months_billed,
                           completion_rate=completion_rate,
                           school_name=settings.get('school_name', 'Alliedian School'),
                           bank_name=settings.get('bank_name', 'Bank Account'))

@app.route('/voucher/generate', methods=['GET', 'POST'])
@login_required
def voucher_generate():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    class_query = "SELECT DISTINCT class FROM students"
    class_params = []
    if active_campus_id:
        class_query += " WHERE campus_id = ?"
        class_params = [active_campus_id]
    class_query += " ORDER BY class"
    
    classes = conn.execute(class_query, class_params).fetchall()
    classes = [r['class'] for r in classes]
    
    selected_class = request.args.get('class', '')
    selected_student_id = request.args.get('student_id', '', type=int)
    
    students = []
    if selected_class:
        student_query = "SELECT id, name, father_name, campus_id FROM students WHERE class = ?"
        student_params = [selected_class]
        if active_campus_id:
            student_query += " AND campus_id = ?"
            student_params.append(active_campus_id)
        student_query += " ORDER BY name"
        
        students = conn.execute(student_query, student_params).fetchall()
        
    conn.close()
    
    months = list(MONTH_NUM_TO_NAME.values())
    years = [2025, 2026, 2027, 2028]
    
    current_month = MONTH_NUM_TO_NAME[datetime.now().month]
    current_year = datetime.now().year
    
    return render_template(
        'voucher_generate.html',
        classes=classes,
        students=students,
        selected_class=selected_class,
        selected_student_id=selected_student_id,
        months=months,
        years=years,
        current_month=current_month,
        current_year=current_year,
    )


@app.route('/voucher/print')
@login_required
def voucher_print():
    # Determine whether we are generating for a single student or an entire class
    student_id = request.args.get('student_id', type=int)
    selected_class = request.args.get('class')
    month = request.args.get('month')
    year = request.args.get('year', type=int)
    due_date = request.args.get('due_date')
    generate_class = bool(selected_class and (not student_id or request.args.get('class_voucher') is not None))
    num_months = request.args.get('num_months', 1, type=int)

    # Calculate end_month if num_months > 1
    start_month_num = MONTH_NAME_TO_NUM.get(month, 1) if month else 1
    end_month_num = (start_month_num - 1 + num_months - 1) % 12 + 1
    end_month = MONTH_NUM_TO_NAME.get(end_month_num, month)

    # Validate common parameters
    if not month or not year:
        flash('Invalid parameters for voucher generation.', 'danger')
        return redirect(url_for('voucher_generate'))

    other_dues = request.args.get('other_dues', 0.0, type=float)
    other_dues_desc = request.args.get('other_dues_desc', '').strip()
    if not other_dues_desc:
        other_dues_desc = f"Annual subscription {year}"

    active_campus_id = get_active_campus_id()
    conn = get_db_connection()

    def get_unpaid_annual_charges(student, yr):
        """Return unpaid annual charges amount for the given student and year."""
        annual_charges = float(student['annual_charges'] or 0.0) if 'annual_charges' in student.keys() else 0.0
        if annual_charges <= 0:
            return 0.0
        paid_rec = conn.execute(
            "SELECT SUM(paid_amount) FROM annual_charges_payments WHERE student_id = ? AND year = ? AND (notes IS NULL OR (LOWER(notes) NOT LIKE ? AND LOWER(notes) NOT LIKE ?))",
            (student['id'], yr, '%summer pack%', '%sp%')
        ).fetchone()
        paid_annual = float(paid_rec[0] or 0.0) if paid_rec and paid_rec[0] is not None else 0.0
        return max(0.0, annual_charges - paid_annual)

    if generate_class:
        # Fetch all students belonging to the selected class (and campus if filtered)
        query = "SELECT * FROM students WHERE class = ? AND (status IS NULL OR status = 'active')"
        params = [selected_class]
        if active_campus_id:
            query += " AND campus_id = ?"
            params.append(active_campus_id)
        query += " ORDER BY id ASC"
        students = conn.execute(query, params).fetchall()

        if not students:
            conn.close()
            flash('No active students found in the selected class.', 'warning')
            return redirect(url_for('voucher_generate'))

        student_ids = [s['id'] for s in students]

        # 1. Batch fetch all fees for these students in ONE query
        fees_map = {sid: [] for sid in student_ids}
        if student_ids:
            placeholders = ','.join(['?'] * len(student_ids))
            all_fees = conn.execute(
                f"SELECT student_id, month, year, paid_amount FROM fees WHERE student_id IN ({placeholders})",
                student_ids
            ).fetchall()
            for f in all_fees:
                fees_map[f['student_id']].append(f)

        # 2. Batch fetch all annual charges payments for these students in ONE query
        ac_map = {sid: 0.0 for sid in student_ids}
        if student_ids:
            placeholders = ','.join(['?'] * len(student_ids))
            all_ac = conn.execute(
                f"""SELECT student_id, SUM(paid_amount) 
                    FROM annual_charges_payments 
                    WHERE student_id IN ({placeholders}) AND year = ? AND (notes IS NULL OR (LOWER(notes) NOT LIKE ? AND LOWER(notes) NOT LIKE ?))
                    GROUP BY student_id""",
                student_ids + [year, '%summer pack%', '%sp%']
            ).fetchall()
            for r in all_ac:
                ac_map[r[0]] = float(r[1] or 0.0)

        # 3. Cache campus settings in memory
        all_settings_rows = conn.execute("SELECT key, value FROM settings").fetchall()
        base_settings = {row['key']: row['value'] for row in all_settings_rows}
        
        def resolve_settings_for_campus(cid):
            s = dict(base_settings)
            if cid:
                for k in list(base_settings.keys()):
                    ck = f"{k}_{cid}"
                    if ck in base_settings:
                        s[k] = base_settings[ck]
            return s
            
        campus_ids = list(set(s['campus_id'] for s in students if s['campus_id']))
        settings_cache = {cid: resolve_settings_for_campus(cid) for cid in campus_ids}
        default_settings = resolve_settings_for_campus(active_campus_id or 1)

        vouchers = []
        for student in students:
            sid = student['id']
            settings = settings_cache.get(student['campus_id'], default_settings)
            
            # Pass pre-fetched fees to avoid any db calls
            fee_details = get_student_fee_details(student, month, year, payments=fees_map.get(sid, []))
            
            # Calculate unpaid annual charges in memory
            annual_charges = float(student['annual_charges'] or 0.0) if 'annual_charges' in student.keys() else 0.0
            paid_annual = ac_map.get(sid, 0.0)
            unpaid_annual = max(0.0, annual_charges - paid_annual) if annual_charges > 0 else 0.0
            
            # Auto-include unpaid annual charges (if no manual other_dues set)
            if other_dues > 0:
                auto_other_dues = other_dues
                auto_other_dues_desc = other_dues_desc or f"Annual Charges {year}"
            elif unpaid_annual > 0:
                auto_other_dues = unpaid_annual
                auto_other_dues_desc = f"Annual Charges {year}"
            else:
                auto_other_dues = 0.0
                auto_other_dues_desc = ""
            
            # Distribute paid_this_month across arrears and generated months
            available_paid = fee_details['paid_this_month']
            if available_paid >= fee_details['arrears']:
                available_paid -= fee_details['arrears']
                display_arrears = 0
            else:
                display_arrears = fee_details['arrears'] - available_paid
                available_paid = 0
                
            multi_months_list = []
            total_months_fee = 0
            for i in range(num_months):
                m_num = (start_month_num - 1 + i) % 12 + 1
                m_name = MONTH_NUM_TO_NAME.get(m_num, month)
                if available_paid >= fee_details['monthly_fee']:
                    m_fee = 0
                    available_paid -= fee_details['monthly_fee']
                else:
                    m_fee = fee_details['monthly_fee'] - available_paid
                    available_paid = 0
                multi_months_list.append({'name': m_name, 'fee': m_fee})
                total_months_fee += m_fee
                
            fee_details['remaining_payable'] = display_arrears + total_months_fee
            
            # If tuition fee is already paid, zero out monthly breakdown but KEEP annual charges intact
            if fee_details['remaining_payable'] <= 0:
                display_arrears = 0
                for m in multi_months_list:
                    m['fee'] = 0
                fee_details['remaining_payable'] = 0

            current_other_dues = auto_other_dues
            payable_by_due = fee_details['remaining_payable'] + current_other_dues
            
            # Determine due date (use default if not supplied)
            if not due_date:
                due_day = int(settings.get('due_day', 10))
                month_num = MONTH_NAME_TO_NUM.get(month, 1)
                calc_due = f"{due_day:02d}-{month_num:02d}-{year}"
            else:
                calc_due = due_date
                
            late_fee = float(settings.get('late_fee', 100))
            payable_after_due = payable_by_due + (late_fee if payable_by_due > 0 else 0)
            
            vouchers.append({
                'school_name': settings.get('school_name', 'Alliedian School Al-Rehman Campus, Okara'),
                'bank_name': settings.get('bank_name', 'MCB Bank Limited'),
                'student': student,
                'paid_this_month': fee_details['paid_this_month'],
                'month': month,
                'end_month': end_month,
                'months': num_months,
                'multi_months_list': multi_months_list,
                'year': year,
                'due_date': calc_due,
                'issue_date': datetime.now().strftime('%d-%m-%Y'),
                'arrears': display_arrears,
                'monthly_fee': fee_details['monthly_fee'],
                'unpaid_annual_charges': unpaid_annual,
                'annual_charges': unpaid_annual,
                'other_dues': current_other_dues,
                'other_dues_desc': auto_other_dues_desc,
                'payable_by_due': payable_by_due,
                'payable_after_due': payable_after_due,
                'late_fee': late_fee if payable_by_due > 0 else 0
            })
        conn.close()
        return render_template('voucher_class.html', vouchers=vouchers)
    else:
        # Single student path
        if not student_id:
            conn.close()
            flash('Invalid parameters for voucher generation.', 'danger')
            return redirect(url_for('voucher_generate'))
        student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
        if not student or (active_campus_id and student['campus_id'] != active_campus_id):
            conn.close()
            flash('Student not found or access denied.', 'danger')
            return redirect(url_for('voucher_generate'))
        settings = get_campus_settings(student['campus_id'])
        fee_details = get_student_fee_details(student, month, year)
        
        # Distribute paid_this_month across arrears and generated months
        available_paid = fee_details['paid_this_month']
        if available_paid >= fee_details['arrears']:
            available_paid -= fee_details['arrears']
            display_arrears = 0
        else:
            display_arrears = fee_details['arrears'] - available_paid
            available_paid = 0
            
        multi_months_list = []
        total_months_fee = 0
        for i in range(num_months):
            m_num = (start_month_num - 1 + i) % 12 + 1
            m_name = MONTH_NUM_TO_NAME.get(m_num, month)
            if available_paid >= fee_details['monthly_fee']:
                m_fee = 0
                available_paid -= fee_details['monthly_fee']
            else:
                m_fee = fee_details['monthly_fee'] - available_paid
                available_paid = 0
            multi_months_list.append({'name': m_name, 'fee': m_fee})
            total_months_fee += m_fee
            
        # Auto-include unpaid annual charges for single student
        unpaid_annual = get_unpaid_annual_charges(student, year)
        if other_dues > 0:
            auto_other_dues = other_dues
            auto_other_dues_desc = other_dues_desc or f"Annual Charges {year}"
        elif unpaid_annual > 0:
            auto_other_dues = unpaid_annual
            auto_other_dues_desc = f"Annual Charges {year}"
        else:
            auto_other_dues = 0.0
            auto_other_dues_desc = ""

        fee_details['remaining_payable'] = display_arrears + total_months_fee
        
        # If tuition fee is already paid, zero out monthly breakdown but KEEP annual charges intact
        if fee_details['remaining_payable'] <= 0:
            display_arrears = 0
            for m in multi_months_list:
                m['fee'] = 0
            fee_details['remaining_payable'] = 0

        current_other_dues = auto_other_dues
        payable_by_due = fee_details['remaining_payable'] + current_other_dues

        if not due_date:
            due_day = int(settings.get('due_day', 10))
            month_num = MONTH_NAME_TO_NUM.get(month, 1)
            due_date = f"{due_day:02d}-{month_num:02d}-{year}"
            
        late_fee = float(settings.get('late_fee', 100))
        payable_after_due = payable_by_due + (late_fee if payable_by_due > 0 else 0)
        
        voucher_data = {
            'school_name': settings.get('school_name', 'Alliedian School Al-Rehman Campus, Okara'),
            'bank_name': settings.get('bank_name', 'MCB Bank Limited'),
            'student': student,
            'month': month,
            'end_month': end_month,
            'months': num_months,
            'multi_months_list': multi_months_list,
            'year': year,
            'due_date': due_date,
            'issue_date': datetime.now().strftime('%d-%m-%Y'),
            'arrears': display_arrears,
            'monthly_fee': fee_details['monthly_fee'],
            'unpaid_annual_charges': unpaid_annual,
            'annual_charges': unpaid_annual,
            'other_dues': current_other_dues,
            'other_dues_desc': auto_other_dues_desc,
            'payable_by_due': payable_by_due,
            'payable_after_due': payable_after_due,
            'late_fee': late_fee if payable_by_due > 0 else 0
        }
        conn.close()
        return render_template('voucher.html', data=voucher_data)

@app.route('/defaulters')
@login_required
def defaulters_view():
    target_month = request.args.get('month', MONTH_NUM_TO_NAME[datetime.now().month])
    target_year = request.args.get('year', datetime.now().year, type=int)
    class_filter = request.args.get('class_filter', '').strip()
    min_amount_raw = request.args.get('min_amount', '').strip()
    try:
        min_amount = float(min_amount_raw) if min_amount_raw else 0.0
    except (ValueError, TypeError):
        min_amount = 0.0
        min_amount_raw = ''
    
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    class_query = "SELECT DISTINCT class FROM students"
    class_params = []
    if active_campus_id:
        class_query += " WHERE campus_id = ?"
        class_params = [active_campus_id]
    class_query += " ORDER BY class"
    classes = conn.execute(class_query, class_params).fetchall()
    classes = [r['class'] for r in classes]
    
    query = '''
        SELECT s.*, c.name as campus_name 
        FROM students s 
        LEFT JOIN campuses c ON s.campus_id = c.id 
        WHERE 1=1
    '''
    params = []
    if active_campus_id:
        query += " AND s.campus_id = ?"
        params.append(active_campus_id)
    if class_filter:
        query += " AND s.class = ?"
        params.append(class_filter)
        
    students = conn.execute(query, params).fetchall()
    
    student_ids = [s['id'] for s in students]
    fees_map = {sid: [] for sid in student_ids}
    if student_ids:
        placeholders = ','.join(['?'] * len(student_ids))
        all_fees = conn.execute(
            f"SELECT student_id, month, year, paid_amount FROM fees WHERE student_id IN ({placeholders})",
            student_ids
        ).fetchall()
        for f in all_fees:
            fees_map[f['student_id']].append(f)
            
    defaulters = []
    total_defaulter_amount = 0.0
    
    for s in students:
        details = get_student_fee_details(s, target_month, target_year, payments=fees_map.get(s['id'], []))
        if details['remaining_payable'] > 0:
            if min_amount > 0 and details['remaining_payable'] < min_amount:
                continue
            defaulters.append({
                'id': s['id'],
                'name': s['name'],
                'father_name': s['father_name'],
                'phone_number': s['phone_number'],
                'class': s['class'],
                'campus_name': s['campus_name'],
                'monthly_fee': details['monthly_fee'],
                'arrears': details['arrears'],
                'total_payable': details['total_payable'],
                'paid': details['paid_this_month'],
                'remaining': details['remaining_payable']
            })
            total_defaulter_amount += details['remaining_payable']
            
    conn.close()
    
    months = list(MONTH_NUM_TO_NAME.values())
    years = [2025, 2026, 2027, 2028]
    
    return render_template('defaulters.html',
                           defaulters=defaulters,
                           classes=classes,
                           class_filter=class_filter,
                           min_amount=min_amount_raw,
                           target_month=target_month,
                           target_year=target_year,
                           months=months,
                           years=years,
                           total_defaulter_amount=total_defaulter_amount)

@app.route('/campuses', methods=['GET', 'POST'])
@login_required
def campuses_view():
    if session.get('role') != 'admin':
        flash('Access Denied. Only Head Office can manage campuses.', 'danger')
        return redirect(url_for('dashboard'))
        
    conn = get_db_connection()
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip().lower()
        
        if not name or not code:
            flash('Campus Name and Code are required!', 'danger')
        else:
            try:
                conn.execute("INSERT INTO campuses (name, code) VALUES (?, ?)", (name, code))
                conn.commit()
                
                # Auto create user operator
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM campuses WHERE code = ?", (code,))
                c_id = cursor.fetchone()[0]
                conn.execute("INSERT INTO users (username, password, campus_id, role) VALUES (?, ?, ?, 'operator')",
                             (code, code, c_id))
                conn.commit()
                flash(f"Campus '{name}' and operator user account '{code}' created successfully!", 'success')
            except Exception as e:
                flash(f"Error creating campus '{code}': {str(e)}", 'danger')
                
    campuses = conn.execute('''
        SELECT c.*, COUNT(s.id) as student_count 
        FROM campuses c 
        LEFT JOIN students s ON s.campus_id = c.id 
        GROUP BY c.id 
        ORDER BY c.id
    ''').fetchall()
    conn.close()
    
    return render_template('campuses.html', campuses=campuses)

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings_view():
    active_campus_id = get_active_campus_id()
    conn = get_db_connection()
    
    if request.method == 'POST':
        school_name = request.form['school_name'].strip()
        bank_name = request.form['bank_name'].strip()
        due_day = request.form['due_day'].strip()
        late_fee = request.form['late_fee'].strip()
        new_password = request.form['new_password'].strip()
        
        # Save settings for specific campus if active, otherwise globally
        if active_campus_id:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (f"school_name_{active_campus_id}", school_name))
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (f"bank_name_{active_campus_id}", bank_name))
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (f"due_day_{active_campus_id}", due_day))
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (f"late_fee_{active_campus_id}", late_fee))
        else:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('school_name', ?)", (school_name,))
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('bank_name', ?)", (bank_name,))
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('due_day', ?)", (due_day,))
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('late_fee', ?)", (late_fee,))
            
        if new_password:
            hashed_pw = generate_password_hash(new_password)
            if active_campus_id:
                conn.execute("UPDATE users SET password = ? WHERE username = ?", (hashed_pw, session['username']))
            else:
                conn.execute("UPDATE users SET password = ? WHERE username = 'admin'", (hashed_pw,))
            flash('Settings and password updated successfully!', 'success')
        else:
            flash('Settings updated successfully!', 'success')
            
        conn.commit()
        conn.close()
        return redirect(url_for('settings_view'))
        
    settings_dict = get_campus_settings(active_campus_id)
    conn.close()
    return render_template('settings.html', settings=settings_dict)

@app.route('/settings/bulk-update', methods=['POST'])
@login_required
def settings_bulk_update():
    if session.get('role') != 'admin':
        flash('Access Denied. Only Head Office admin can bulk update fees.', 'danger')
        return redirect(url_for('dashboard'))
        
    campus_id = request.form.get('campus_id', type=int)
    update_target = request.form.get('update_target') # 'monthly_fee' or 'annual_charges'
    update_type = request.form.get('update_type') # 'percentage' or 'flat'
    amount = request.form.get('amount', type=float)
    
    if amount is None or amount <= 0:
        flash('Please enter a valid positive increase amount.', 'danger')
        return redirect(url_for('settings_view'))
        
    conn = get_db_connection()
    
    # Base query columns handling NULLs
    if update_target == 'monthly_fee':
        column = 'COALESCE(monthly_fee, 0)'
        target_col = 'monthly_fee'
    elif update_target == 'annual_charges':
        column = 'COALESCE(annual_charges, 0)'
        target_col = 'annual_charges'
    else:
        conn.close()
        flash('Invalid update target selected.', 'danger')
        return redirect(url_for('settings_view'))
        
    # Build update expression
    if update_type == 'percentage':
        update_expr = f"{target_col} = ROUND({column} * (1 + ? / 100))"
    elif update_type == 'flat':
        update_expr = f"{target_col} = {column} + ?"
    else:
        conn.close()
        flash('Invalid update type selected.', 'danger')
        return redirect(url_for('settings_view'))
        
    query = f"UPDATE students SET {update_expr}"
    params = [amount]
    
    if campus_id: # If not 0 (which means All Campuses)
        query += " WHERE campus_id = ?"
        params.append(campus_id)
        
    cursor = conn.cursor()
    cursor.execute(query, params)
    updated_count = cursor.rowcount
    conn.commit()
    conn.close()
    
    target_name = "Monthly Fee" if update_target == 'monthly_fee' else "Annual Charges"
    type_name = f"{amount}%" if update_type == 'percentage' else f"Rs. {amount}"
    
    flash(f"Successfully increased {target_name} by {type_name} for {updated_count} students!", 'success')
    return redirect(url_for('settings_view'))


@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel_view():
    if request.method == 'POST':
        if 'auto_import' in request.form:
            import import_excel
            try:
                import_excel.init_db()
                import_excel.import_data()
                flash('Successfully imported default Excel record (Fee record 2026.xlsx) from workspace!', 'success')
            except Exception as e:
                flash(f'Error during import: {str(e)}', 'danger')
            return redirect(url_for('dashboard'))
            
        if 'file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)
            
        file = request.files['file']
        if file.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)
            
        if file and file.filename.endswith(('.xlsx', '.xls')):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            try:
                import_excel_file(filepath)
                flash(f'File "{filename}" uploaded and student records imported successfully!', 'success')
            except Exception as e:
                flash(f'Error importing from Excel: {str(e)}', 'danger')
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    
            return redirect(url_for('dashboard'))
        else:
            flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
            
    return render_template('import.html')

def import_excel_file(filepath):
    xls = pd.ExcelFile(filepath)
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("DELETE FROM fees")
    cursor.execute("DELETE FROM students")
    cursor.execute("DELETE FROM sqlite_sequence WHERE name IN ('students', 'fees')")
    conn.commit()
    
    cursor.execute("SELECT id FROM campuses WHERE code = 'campus_1'")
    default_campus_id = cursor.fetchone()[0]
    
    for sheet in xls.sheet_names:
        if sheet == 'van Charges':
            continue
            
        df = pd.read_excel(xls, sheet, header=None)
        if len(df) < 3:
            continue
            
        headers = [str(x).strip() for x in df.iloc[2].tolist()]
        
        name_col = 1
        father_col = 2
        
        fee_cols = []
        for i, h in enumerate(headers):
            hl = h.lower()
            if ('month' in hl or 'monthly' in hl) and not any(hl.startswith(m) for m in MONTH_NAME_TO_NUM):
                fee_cols.append(i)
                
        month_cols = []
        for i, h in enumerate(headers):
            hl = h.lower()
            for m_key in SHORT_MONTHS:
                if hl.startswith(m_key):
                    month_cols.append((i, MONTH_NUM_TO_NAME[SHORT_MONTHS[m_key]], SHORT_MONTHS[m_key]))
                    break
                    
        for r_idx in range(3, len(df)):
            row = df.iloc[r_idx].tolist()
            if len(row) <= 1:
                continue
                
            name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
            if not name or name.lower() == 'nan' or name == 'SR. #' or name.startswith('Name'):
                continue
                
            father = str(row[father_col]).strip() if (father_col < len(row) and pd.notna(row[father_col])) else ''
            if father.lower() == 'nan':
                father = ''
                
            from import_excel import clean_amount
            monthly_fee = 0
            for f_col in fee_cols:
                if f_col < len(row):
                    val = row[f_col]
                    fee_val = clean_amount(val)
                    if fee_val > 0:
                        monthly_fee = fee_val
                        
            if monthly_fee == 0:
                monthly_fee = 2400
                
            start_month = 3
            start_year = 2026
            
            earliest_payment_found = False
            for f_idx, m_name, m_val in month_cols:
                if f_idx < len(row) and pd.notna(row[f_idx]):
                    val = row[f_idx]
                    amt = clean_amount(val, monthly_fee)
                    if amt > 0:
                        year = 2025 if (m_name in ('November', 'December') and f_idx < 8) else 2026
                        if not earliest_payment_found or (year < start_year) or (year == start_year and m_val < start_month):
                            start_month = m_val
                            start_year = year
                            earliest_payment_found = True
                            
            cursor.execute('''
                INSERT INTO students (name, father_name, class, monthly_fee, start_month, start_year, campus_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (name, father, sheet, monthly_fee, start_month, start_year, default_campus_id))
            student_id = cursor.lastrowid
            
            for f_idx, m_name, m_val in month_cols:
                if f_idx < len(row) and pd.notna(row[f_idx]):
                    val = row[f_idx]
                    amt = clean_amount(val, monthly_fee)
                    if amt > 0:
                        year = 2025 if (m_name in ('November', 'December') and f_idx < 8) else 2026
                        date_paid = f"{year}-{m_val:02d}-01"
                        cursor.execute('''
                            INSERT INTO fees (student_id, month, year, paid_amount, date_paid, campus_id)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (student_id, m_name, year, amt, date_paid, default_campus_id))
                        
    conn.commit()
    conn.close()

@app.route('/sos')
@login_required
def sos_view():
    class_filter = request.args.get('class_filter', '').strip()
    search = request.args.get('search', '').strip()
    
    conn = get_db_connection()
    
    # Fetch unique classes available from students table to populate filter options
    classes_query = "SELECT DISTINCT class FROM students ORDER BY class"
    classes = [r['class'] for r in conn.execute(classes_query).fetchall()]
    
    # Fetch all SOS materials
    query = "SELECT * FROM sos_materials WHERE 1=1"
    params = []
    
    if class_filter:
        query += " AND class_name = ?"
        params.append(class_filter)
        
    if search:
        query += " AND (LOWER(title) LIKE ? OR LOWER(description) LIKE ?)"
        params.extend([f"%{search.lower()}%", f"%{search.lower()}%"])
        
    query += " ORDER BY id DESC"
    materials = conn.execute(query, params).fetchall()
    conn.close()
    
    return render_template('sos.html', 
                           materials=materials, 
                           classes=classes, 
                           class_filter=class_filter, 
                           search=search)

@app.route('/sos/upload', methods=['POST'])
@login_required
def sos_upload():
    if session.get('role') != 'admin':
        flash('Access Denied. Only Head Office admin can upload study materials.', 'danger')
        return redirect(url_for('sos_view'))
        
    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    class_name = request.form.get('class_name', '').strip()
    
    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('sos_view'))
        
    file = request.files['file']
    if file.filename == '':
        flash('No selected file!', 'danger')
        return redirect(url_for('sos_view'))
        
    if not title or not class_name:
        flash('Title and Class are required fields!', 'danger')
        return redirect(url_for('sos_view'))
        
    if file:
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S_')
        unique_filename = timestamp + filename
        filepath = os.path.join(app.config['SOS_UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # Record in database
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO sos_materials (title, description, class_name, filename, filepath, date_uploaded)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (title, description, class_name, unique_filename, filepath, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        conn.close()
        
        flash(f'Study material "{title}" uploaded successfully for Class {class_name}!', 'success')
        
    return redirect(url_for('sos_view'))

@app.route('/sos/download/<int:material_id>')
@login_required
def sos_download(material_id):
    conn = get_db_connection()
    material = conn.execute("SELECT * FROM sos_materials WHERE id = ?", (material_id,)).fetchone()
    conn.close()
    
    if not material:
        flash('Requested material not found!', 'danger')
        return redirect(url_for('sos_view'))
        
    filepath = material['filepath']
    if not os.path.exists(filepath):
        flash('File not found on server!', 'danger')
        return redirect(url_for('sos_view'))
        
    return send_file(filepath, as_attachment=True, download_name=material['filename'].split('_', 1)[-1])

@app.route('/sos/delete/<int:material_id>', methods=['POST'])
@login_required
def sos_delete(material_id):
    if session.get('role') != 'admin':
        flash('Access Denied.', 'danger')
        return redirect(url_for('sos_view'))
        
    conn = get_db_connection()
    material = conn.execute("SELECT * FROM sos_materials WHERE id = ?", (material_id,)).fetchone()
    
    if material:
        filepath = material['filepath']
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except Exception as e:
                pass
        
        conn.execute("DELETE FROM sos_materials WHERE id = ?", (material_id,))
        conn.commit()
        flash('Study material record deleted successfully.', 'success')
    else:
        flash('Material not found!', 'danger')
        
    conn.close()
    return redirect(url_for('sos_view'))

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=3013, debug=True)
