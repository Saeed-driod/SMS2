import os
import re
import openpyxl
from db import get_db_connection, is_postgres, init_db

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'Fee record Main Campus.xlsx')

MONTH_LOOKUP = {
    'jan': ('January', 1), 'january': ('January', 1),
    'feb': ('February', 2), 'fe': ('February', 2), 'february': ('February', 2),
    'mar': ('March', 3), 'march': ('March', 3),
    'apr': ('April', 4), 'april': ('April', 4),
    'may': ('May', 5),
    'jun': ('June', 6), 'june': ('June', 6),
    'jul': ('July', 7), 'july': ('July', 7),
    'aug': ('August', 8), 'august': ('August', 8),
    'sep': ('September', 9), 'sept': ('September', 9), 'september': ('September', 9),
    'oct': ('October', 10), 'october': ('October', 10),
    'nov': ('November', 11), 'november': ('November', 11),
    'dec': ('December', 12), 'dece': ('December', 12), 'december': ('December', 12)
}

MONTH_NUM_TAG = {
    'jan': 1, 'january': 1,
    'feb': 2, 'fe': 2, 'february': 2,
    'mar': 3, 'march': 3,
    'apr': 4, 'april': 4,
    'may': 5,
    'jun': 6, 'june': 6,
    'jul': 7, 'july': 7,
    'aug': 8, 'august': 8,
    'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10,
    'nov': 11, 'november': 11,
    'dec': 12, 'dece': 12, 'december': 12
}

def derive_start_month_year(arrears_tag, default_start_m=3, default_start_y=2026):
    if not arrears_tag:
        return default_start_m, default_start_y
    tag = str(arrears_tag).lower().strip()
    
    # Find matching month
    matched_m = None
    for mk, mnum in MONTH_NUM_TAG.items():
        if mk in tag:
            matched_m = mnum
            break
            
    if not matched_m:
        return default_start_m, default_start_y
        
    if matched_m == 11:
        return 12, 2025 # next month is Dec 2025
    elif matched_m == 12:
        return 1, 2026 # next month is Jan 2026
    elif matched_m in (1, 2):
        return matched_m + 1, 2026
    elif matched_m >= 3 and matched_m <= 10:
        return matched_m + 1, 2026
        
    return default_start_m, default_start_y

def parse_pending_amount(val):
    """Extracts pending amount from strings like '2000p', '4500(2000p)', '4000(1000p', '1500(3500p)'"""
    if val is None:
        return 0
    s = str(val).strip().lower()
    m_split = re.search(r'\((\d+)\s*p?\)?', s)
    if m_split:
        return int(m_split.group(1))
    m_p = re.search(r'^(\d+)\s*p$', s)
    if m_p:
        return int(m_p.group(1))
    return 0

def is_month_arrears(val_str):
    """Checks if value is like '21800july', '60600jul', '4800nov', '15200pending till june'"""
    val_lower = val_str.lower().strip()
    m = re.match(r'^(\d+)\s*([a-zA-Z]+.*)$', val_lower)
    if m:
        amt = int(m.group(1))
        suffix = m.group(2).strip().lower()
        if suffix not in ('p', 'st', 'stat', 'party', 'prty') and any(m_key in suffix for m_key in ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec','clear','pend']):
            return True, amt, suffix
    if 'pending till' in val_lower or 'till' in val_lower:
        m_amt = re.search(r'\d+', val_str)
        if m_amt:
            return True, int(m_amt.group()), val_str
    return False, 0, ''

def parse_cell_amount(val, default_fee=0):
    if val is None or val == '':
        return 0, ''
    if isinstance(val, (int, float)):
        if val == 0:
            return default_fee, 'Waived (0)'
        return int(val), ''
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('nan', 'none', '-'):
        return 0, ''
        
    val_lower = val_str.lower()
    
    if val_lower in ('0', '0.0', 'zero', 'nil', 'free', 'waived'):
        return default_fee, 'Waived (0)'
    
    # Check if month arrears e.g. 21800july
    is_arr, arr_amt, arr_suffix = is_month_arrears(val_str)
    if is_arr:
        return 0, f'arrears:{arr_amt}:{arr_suffix}'

    # Check if 'done' or 'paid' or 'clear'
    if any(k in val_lower for k in ('done', 'paid', 'clear', 'ok', 'yes')):
        return default_fee, val_str

    # Pattern like '1500(3500p)' or '4000(4000p)'
    m_split = re.match(r'(\d+)\s*\(\s*(\d+)\s*p?\s*\)', val_lower)
    if m_split:
        paid = int(m_split.group(1))
        pending = int(m_split.group(2))
        return paid, f'{pending} pending'

    # Pattern like '650p' or '2700p' or '5000p'
    m_p_only = re.match(r'^(\d+)\s*p$', val_lower)
    if m_p_only:
        pending = int(m_p_only.group(1))
        return 0, f'{pending} pending'

    # Pattern like '1500st' or '1250party'
    m_num_suffix = re.match(r'^(\d+)\s*([a-zA-Z]+.*)$', val_lower)
    if m_num_suffix:
        amt = int(m_num_suffix.group(1))
        note = m_num_suffix.group(2).strip()
        return amt, note

    # General extract number
    m = re.search(r'\d+', val_str)
    if m:
        return int(m.group()), val_str
        
    return 0, val_str

def parse_ac_cell(val, default_ac=2700):
    if val is None or str(val).strip() == '':
        return default_ac, 0, 'blank_pending'
    val_str = str(val).strip()
    val_lower = val_str.lower()
    
    if val_lower in ('none', 'nan', '-'):
        return default_ac, 0, 'blank_pending'
    if val_lower in ('0', 'zero', 'nil'):
        return 0, 0, 'exempt_zero'
        
    # Pattern like '5000(3500p)' or '1500(3500p)'
    m_split = re.match(r'^(\d+)\s*\(\s*(\d+)\s*p?\s*\)', val_lower)
    if m_split:
        p1 = int(m_split.group(1))
        p2 = int(m_split.group(2))
        if p1 > p2:
            total_ac = p1
            paid_ac = p1 - p2
        else:
            paid_ac = p1
            total_ac = p1 + p2
        return total_ac, paid_ac, f'{p2} pending'

    # Pattern like '2700p' or '1350p' or '5000p'
    m_p_only = re.match(r'^(\d+)\s*p$', val_lower)
    if m_p_only:
        pending = int(m_p_only.group(1))
        return pending, 0, f'{pending} pending'

    # Number with party like '1250party'
    m_party = re.match(r'^(\d+)\s*part', val_lower)
    if m_party:
        paid = int(m_party.group(1))
        return paid, paid, 'partial'

    # Pure number like 2700
    m_num = re.search(r'\d+', val_str)
    if m_num:
        amt = int(m_num.group())
        return amt, amt, 'paid'
        
    return default_ac, 0, val_str

def is_row_red(ws, r):
    for c in range(1, min(6, ws.max_column+1)):
        cell = ws.cell(r, c)
        fill = cell.fill
        if fill and fill.fill_type:
            fg = fill.fgColor
            rgb = str(getattr(fg, 'rgb', ''))
            theme = getattr(fg, 'theme', None)
            tint = getattr(fg, 'tint', 0.0)
            if rgb in ('FFFF0000', 'FFC00000', 'FFD99795', 'FFFF6666', 'FFFFC7CE', 'FFFF2020') or (theme == 9 and tint < 0):
                return True
    return False

def import_main_campus():
    init_db()
    
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: {EXCEL_PATH} not found!")
        return
        
    print(f"Loading Excel file: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # 1. Identify Main Campus ID
    campus_row = cur.execute("SELECT id, name FROM campuses WHERE code = 'main_campus' OR name LIKE '%Main%' ORDER BY id ASC LIMIT 1").fetchone()
    if campus_row:
        campus_id = campus_row['id']
        campus_name = campus_row['name']
    else:
        campus_row = cur.execute("SELECT id, name FROM campuses ORDER BY id ASC LIMIT 1").fetchone()
        campus_id = campus_row['id'] if campus_row else 1
        campus_name = campus_row['name'] if campus_row else 'Default Campus'
        
    print(f"Importing to Campus: {campus_name} (ID: {campus_id})")
    
    # 2. Clear existing records for this campus to prevent duplicates
    cur.execute("DELETE FROM fees WHERE campus_id = ?", (campus_id,))
    cur.execute("DELETE FROM annual_charges_payments WHERE campus_id = ?", (campus_id,))
    cur.execute("DELETE FROM students WHERE campus_id = ?", (campus_id,))
    conn.commit()
    print("Previous records for this campus cleared successfully.")
    
    total_students_imported = 0
    total_active = 0
    total_withdrawn = 0
    total_arrears_students = 0
    total_arrears_amount = 0
    total_fee_records = 0
    total_fee_amount = 0
    total_ac_records = 0
    total_ac_amount = 0
    
    for sheetname in wb.sheetnames:
        if sheetname == 'van Charges':
            continue
            
        ws = wb[sheetname]
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(3, c).value
            headers.append(str(v).strip() if v is not None else '')
            
        # Check AC / Annual columns
        ac_cols = []
        for c in range(4, min(14, ws.max_column + 1)):
            h = headers[c-1].lower() if c-1 < len(headers) else ''
            if h in ('ac', 'sp', 'annual'):
                ac_cols.append((c, headers[c-1]))
                
        sheet_students_data = []
        sheet_ac_data = []
        sheet_fee_data = []
        
        for r in range(4, ws.max_row + 1):
            name_val = ws.cell(r, 2).value
            if not name_val or str(name_val).strip().lower() in ('none', 'nan', '', 'student name', 'name', 'sr. #', 'sr#'):
                continue
                
            name = str(name_val).strip()
            father = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else ''
            if father.lower() in ('none', 'nan'):
                father = ''
                
            status = 'withdrawn' if is_row_red(ws, r) else 'active'
            if status == 'withdrawn':
                total_withdrawn += 1
            else:
                total_active += 1
                
            # Determine monthly fee
            monthly_fee = 0
            for c in range(4, min(14, ws.max_column + 1)):
                h = headers[c-1].lower() if c-1 < len(headers) else ''
                if ('month' in h or 'monthly' in h) and not any(h.startswith(m) for m in MONTH_LOOKUP):
                    v = ws.cell(r, c).value
                    amt, _ = parse_cell_amount(v)
                    if amt > 0:
                        monthly_fee = amt
            if monthly_fee == 0:
                monthly_fee = 2400
                
            # Calculate opening arrears from month-tagged entries (like 21800july or 20700may)
            student_opening_arrears = 0
            arrears_tag = None
            for c in range(4, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    val_str = str(v).strip()
                    is_arr, arr_amt, arr_suffix = is_month_arrears(val_str)
                    if is_arr:
                        student_opening_arrears += arr_amt
                        arrears_tag = arr_suffix
                        
            # Also add pending dues from non-monthly columns (e.g. Book 4500(2000p), Admission 2000p, Sationary 5000p, Sp 500p)
            for c in range(4, ws.max_column + 1):
                h = headers[c-1].lower() if c-1 < len(headers) else ''
                if h in ('month', 'monthly', 'month 26', 'month26', 'monthly 24') or any(h.startswith(m) for m in MONTH_LOOKUP):
                    continue
                if h == 'ac':
                    continue # AC pending is handled directly in annual_charges
                v = ws.cell(r, c).value
                p_amt = parse_pending_amount(v)
                if p_amt > 0:
                    student_opening_arrears += p_amt
                        
            if student_opening_arrears > 0:
                total_arrears_students += 1
                total_arrears_amount += student_opening_arrears
                
            # Derive start_month & start_year from arrears tag (e.g. 20700may -> billing starts June 2026)
            start_month, start_year = derive_start_month_year(arrears_tag, default_start_m=3, default_start_y=2026)
            
            # Calculate Annual Charges for this student from AC column
            class_standard_ac = 2600 if sheetname == 'Graduate' else 2700
            student_annual_charges = 0
            for ac_col_idx, ac_col_name in ac_cols:
                if ac_col_name.lower() == 'ac':
                    v = ws.cell(r, ac_col_idx).value
                    total_ac, paid_ac, ac_note = parse_ac_cell(v, class_standard_ac)
                    student_annual_charges = total_ac
                    
            student_idx = len(sheet_students_data)
            sheet_students_data.append((
                name, father, sheetname, monthly_fee, student_annual_charges,
                student_opening_arrears, start_month, start_year, campus_id, status
            ))
            total_students_imported += 1
            
            # Collect Annual Charges Payments (AC / Sp)
            for ac_col_idx, ac_col_name in ac_cols:
                v = ws.cell(r, ac_col_idx).value
                if v is not None:
                    if ac_col_name.lower() == 'ac':
                        total_ac, paid_ac, ac_note = parse_ac_cell(v, class_standard_ac)
                        if paid_ac > 0:
                            date_paid = "2026-03-01"
                            note_text = f"Annual Charges: {ac_note}"
                            sheet_ac_data.append((student_idx, paid_ac, date_paid, campus_id, note_text))
                            total_ac_records += 1
                            total_ac_amount += paid_ac
                    else:
                        # Sp (Summer Pack)
                        amt, note = parse_cell_amount(v)
                        if amt > 0:
                            date_paid = "2026-06-01"
                            note_text = f"{ac_col_name}: {note}" if note else f"{ac_col_name} Payment"
                            sheet_ac_data.append((student_idx, amt, date_paid, campus_id, note_text))
                            total_ac_records += 1
                            total_ac_amount += amt
                        
            # Collect Monthly Fee Payments
            for c in range(4, ws.max_column + 1):
                h = headers[c-1].lower() if c-1 < len(headers) else ''
                m_matched = None
                for mk, (mname, mnum) in MONTH_LOOKUP.items():
                    if h.startswith(mk):
                        m_matched = (mname, mnum)
                        break
                        
                if m_matched:
                    mname, mnum = m_matched
                    v = ws.cell(r, c).value
                    if v is not None and str(v).strip() != '':
                        paid, note = parse_cell_amount(v, monthly_fee)
                        if paid > 0:
                            year = 2025 if (mname in ('November', 'December') and c < 8) else 2026
                            date_paid = f"{year}-{mnum:02d}-01"
                            note_text = note if note else 'Imported from Excel'
                            sheet_fee_data.append((student_idx, mname, year, paid, date_paid, campus_id, note_text))
                            total_fee_records += 1
                            total_fee_amount += paid

        # Fast Batch Insert for this sheet
        if is_postgres():
            from psycopg2.extras import execute_values
            raw_conn = conn._conn
            raw_cur = raw_conn.cursor()
            
            if sheet_students_data:
                res = execute_values(
                    raw_cur,
                    """INSERT INTO students (name, father_name, class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, campus_id, status)
                       VALUES %s RETURNING id""",
                    sheet_students_data,
                    fetch=True
                )
                student_ids = [r[0] for r in res]
                
                if sheet_ac_data:
                    ac_to_insert = [
                        (student_ids[item[0]], 2026, item[1], item[2], item[3], item[4])
                        for item in sheet_ac_data
                    ]
                    execute_values(
                        raw_cur,
                        """INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, campus_id, notes)
                           VALUES %s""",
                        ac_to_insert
                    )
                    
                if sheet_fee_data:
                    fees_to_insert = [
                        (student_ids[item[0]], item[1], item[2], item[3], item[4], item[5], item[6])
                        for item in sheet_fee_data
                    ]
                    execute_values(
                        raw_cur,
                        """INSERT INTO fees (student_id, month, year, paid_amount, date_paid, campus_id, notes)
                           VALUES %s""",
                        fees_to_insert
                    )
            raw_conn.commit()
        else:
            # SQLite insertion
            student_ids = []
            for s_row in sheet_students_data:
                cur.execute('''
                    INSERT INTO students (name, father_name, class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, campus_id, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', s_row)
                student_ids.append(cur.lastrowid)
                
            for item in sheet_ac_data:
                cur.execute('''
                    INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, campus_id, notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (student_ids[item[0]], 2026, item[1], item[2], item[3], item[4]))
                
            for item in sheet_fee_data:
                cur.execute('''
                    INSERT INTO fees (student_id, month, year, paid_amount, date_paid, campus_id, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (student_ids[item[0]], item[1], item[2], item[3], item[4], item[5], item[6]))
            conn.commit()

        print(f"  Processed Sheet '{sheetname}': {len(sheet_students_data)} students, {len(sheet_fee_data)} fee records.")
        conn.commit()
        
    conn.commit()
    conn.close()
    
    print("\n========================================================")
    print("         MAIN CAMPUS IMPORT COMPLETED SUCCESSFULLY!      ")
    print("========================================================")
    print(f"Total Students Imported: {total_students_imported}")
    print(f"   Active Students: {total_active}")
    print(f"   Withdrawn (Red) Students: {total_withdrawn}")
    print(f"Students with Opening Arrears: {total_arrears_students} (Total: Rs. {total_arrears_amount:,})")
    print(f"Total Monthly Fee Transactions: {total_fee_records} (Total: Rs. {total_fee_amount:,})")
    print(f"Total Annual / AC Charges: {total_ac_records} (Total: Rs. {total_ac_amount:,})")
    print("========================================================\n")


EXCEL_28_PATH = os.path.join(BASE_DIR, 'Fee Record 28 Campus.xlsx')

MONTH_28_MAP = {
    'mar': ('March', 3, 2026),
    'march': ('March', 3, 2026),
    'apr': ('April', 4, 2026),
    'april': ('April', 4, 2026),
    'may': ('May', 5, 2026),
    'jun': ('June', 6, 2026),
    'june': ('June', 6, 2026),
    'jul': ('July', 7, 2026),
    'july': ('July', 7, 2026),
    'aug': ('August', 8, 2026),
    'august': ('August', 8, 2026),
    'sep': ('September', 9, 2026),
    'sept': ('September', 9, 2026),
    'september': ('September', 9, 2026),
    'oct': ('October', 10, 2026),
    'october': ('October', 10, 2026),
    'nov': ('November', 11, 2026),
    'november': ('November', 11, 2026),
    'dec': ('December', 12, 2026),
    'december': ('December', 12, 2026),
    'jan': ('January', 1, 2027),
    'january': ('January', 1, 2027),
    'feb': ('February', 2, 2027),
    'february': ('February', 2, 2027)
}

CLASS_28_MAP = {
    'playgroup': 'PG',
    'pg': 'PG',
    'nursery': 'Nursery',
    'prep': 'Prep',
    'one': 'One',
    'two': 'Two',
    'three': 'Three',
    'four': 'Four',
    'five': 'Five',
    'six': 'Six',
    'seven': 'Seven',
    'eight': 'Eight',
    'nine': 'Nine',
    'ten': 'Ten'
}

def import_28_campus(file_path=None):
    init_db()
    
    target_excel = file_path or EXCEL_28_PATH
    if not os.path.exists(target_excel):
        print(f"Error: {target_excel} not found!")
        return
        
    print(f"\n========================================================")
    print(f"       IMPORTING 28 CAMPUS FEE RECORD                    ")
    print(f"========================================================")
    print(f"Loading Excel file: {target_excel}...")
    wb = openpyxl.load_workbook(target_excel, data_only=True)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # 1. Identify 28 Campus ID
    campus_row = cur.execute("SELECT id, name FROM campuses WHERE code = '28_campus' OR name LIKE '%28%' ORDER BY id ASC LIMIT 1").fetchone()
    if campus_row:
        campus_id = campus_row['id']
        campus_name = campus_row['name']
    else:
        cur.execute("INSERT INTO campuses (name, code) VALUES ('28 Campus', '28_campus')")
        conn.commit()
        campus_id = cur.lastrowid
        campus_name = '28 Campus'
        
    print(f"Target Campus: {campus_name} (ID: {campus_id})")
    
    # 2. Clear previous records for 28 Campus to prevent duplicates
    cur.execute("DELETE FROM fees WHERE campus_id = ?", (campus_id,))
    cur.execute("DELETE FROM annual_charges_payments WHERE campus_id = ?", (campus_id,))
    cur.execute("DELETE FROM students WHERE campus_id = ?", (campus_id,))
    conn.commit()
    print("Previous records for 28 Campus cleared successfully.\n")
    
    total_students_imported = 0
    total_active = 0
    total_withdrawn = 0
    total_fee_records = 0
    total_fee_amount = 0
    total_ac_records = 0
    total_ac_amount = 0
    
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        
        # Dynamically locate header row
        header_r = None
        for r in range(1, min(7, ws.max_row + 1)):
            vals = [str(ws.cell(r, c).value).strip().lower() for c in range(1, min(15, ws.max_column + 1)) if ws.cell(r, c).value is not None]
            if any('name' in v for v in vals):
                header_r = r
                break
                
        if not header_r:
            continue
            
        headers = [str(ws.cell(header_r, c).value).strip() if ws.cell(header_r, c).value is not None else '' for c in range(1, ws.max_column + 1)]
        while headers and not headers[-1]:
            headers.pop()
            
        mapped_class = CLASS_28_MAP.get(sheetname.lower().strip(), sheetname)
        
        # Identify columns
        monthly_fee_col = None
        ac_col = None
        statio_col = None
        month_cols = []
        
        for c in range(4, len(headers) + 1):
            h = headers[c-1].lower().strip()
            if 'monthly' in h or h == 'fee' or h == 'monthly fee':
                monthly_fee_col = c
            elif h in ('a/c', 'ac', 'annual', 'annual charges'):
                ac_col = c
            elif 'statio' in h:
                statio_col = c
            else:
                for mk, mtuple in MONTH_28_MAP.items():
                    if h.startswith(mk):
                        month_cols.append((c, headers[c-1], mtuple))
                        break
                        
        sheet_students_data = []
        sheet_ac_data = []
        sheet_fee_data = []
        
        for r in range(header_r + 1, ws.max_row + 1):
            name_val = ws.cell(r, 2).value
            if not name_val or str(name_val).strip().lower() in ('none', 'nan', '', 'total', 'grand total', 'name', 'sr. #', 'sr#'):
                continue
                
            name = str(name_val).strip()
            father_val = ws.cell(r, 3).value
            father = str(father_val).strip() if father_val and str(father_val).strip().lower() not in ('none', 'nan') else ''
            
            is_red = is_row_red(ws, r)
            has_wd_text = False
            
            # 1. Parse Monthly Fee
            m_fee = 0.0
            if monthly_fee_col:
                v = ws.cell(r, monthly_fee_col).value
                if isinstance(v, (int, float)):
                    m_fee = float(v)
                elif v is not None:
                    m = re.search(r'\d+', str(v))
                    if m:
                        m_fee = float(m.group())
            if m_fee == 0.0:
                m_fee = 2200.0 # Standard fee for 28 Campus
                
            # 2. Parse Annual Charges
            ac_val = ws.cell(r, ac_col).value if ac_col else None
            student_ac = 0.0
            ac_paid = 0.0
            ac_note = ''
            if ac_val is not None:
                if isinstance(ac_val, (int, float)):
                    student_ac = float(ac_val)
                    ac_paid = float(ac_val)
                    ac_note = 'Paid'
                else:
                    ac_str = str(ac_val).strip().lower()
                    if any(w in ac_str for w in ['w/d', 'withdraw', 'with drawl', 'main']):
                        has_wd_text = True
                        student_ac = 0.0
                        ac_paid = 0.0
                    else:
                        m = re.search(r'\d+', ac_str)
                        if m:
                            student_ac = float(m.group())
                            ac_paid = float(m.group())
                            ac_note = 'Paid'
            else:
                student_ac = 2200.0
                
            # 3. Parse Monthly Fee Cells
            student_fees = []
            first_active_month = None
            for c, h_name, (mname, mnum, myear) in month_cols:
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != '':
                    if isinstance(v, (int, float)):
                        val_num = float(v)
                        if val_num > 0:
                            if first_active_month is None:
                                first_active_month = (mnum, myear)
                            student_fees.append((mname, myear, val_num, f'{myear}-{mnum:02d}-01', 'Imported from Excel'))
                    else:
                        v_str = str(v).strip().lower()
                        if any(w in v_str for w in ['w/d', 'withdraw', 'with drawl', 'transfer', 'move']):
                            has_wd_text = True
                        else:
                            m = re.search(r'\d+', v_str)
                            if m:
                                amt = float(m.group())
                                if amt > 0:
                                    if first_active_month is None:
                                        first_active_month = (mnum, myear)
                                    student_fees.append((mname, myear, amt, f'{myear}-{mnum:02d}-01', f'Imported from Excel ({v})'))
                                    
            status = 'withdrawn' if (is_red or has_wd_text) else 'active'
            if status == 'withdrawn':
                total_withdrawn += 1
            else:
                total_active += 1
                
            start_month = first_active_month[0] if first_active_month else 3
            start_year = first_active_month[1] if first_active_month else 2026
            
            student_idx = len(sheet_students_data)
            sheet_students_data.append((
                name, father, mapped_class, m_fee, student_ac, 0.0,
                start_month, start_year, campus_id, status
            ))
            total_students_imported += 1
            
            # Annual Charges Payment
            if ac_paid > 0:
                sheet_ac_data.append((student_idx, ac_paid, '2026-03-01', campus_id, f'Annual Charges: {ac_note}'))
                total_ac_records += 1
                total_ac_amount += ac_paid
                
            # Stationery Payment (Playgroup)
            if statio_col:
                st_val = ws.cell(r, statio_col).value
                if isinstance(st_val, (int, float)) and float(st_val) > 0:
                    sheet_ac_data.append((student_idx, float(st_val), '2026-03-01', campus_id, 'Stationery Payment'))
                    total_ac_records += 1
                    total_ac_amount += float(st_val)
                    
            # Monthly Fee Payments
            for mname, myear, amt, date_paid, note in student_fees:
                sheet_fee_data.append((student_idx, mname, myear, amt, date_paid, campus_id, note))
                total_fee_records += 1
                total_fee_amount += amt
                
        # Batch Insert for Sheet
        if is_postgres():
            from psycopg2.extras import execute_values
            raw_conn = conn._conn
            raw_cur = raw_conn.cursor()
            
            if sheet_students_data:
                res = execute_values(
                    raw_cur,
                    """INSERT INTO students (name, father_name, class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, campus_id, status)
                       VALUES %s RETURNING id""",
                    sheet_students_data,
                    fetch=True
                )
                student_ids = [row[0] for row in res]
                
                if sheet_ac_data:
                    ac_to_insert = [
                        (student_ids[item[0]], 2026, item[1], item[2], item[3], item[4])
                        for item in sheet_ac_data
                    ]
                    execute_values(
                        raw_cur,
                        """INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, campus_id, notes)
                           VALUES %s""",
                        ac_to_insert
                    )
                    
                if sheet_fee_data:
                    fees_to_insert = [
                        (student_ids[item[0]], item[1], item[2], item[3], item[4], item[5], item[6])
                        for item in sheet_fee_data
                    ]
                    execute_values(
                        raw_cur,
                        """INSERT INTO fees (student_id, month, year, paid_amount, date_paid, campus_id, notes)
                           VALUES %s""",
                        fees_to_insert
                    )
            raw_conn.commit()
        else:
            student_ids = []
            for s_row in sheet_students_data:
                cur.execute('''
                    INSERT INTO students (name, father_name, class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, campus_id, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', s_row)
                student_ids.append(cur.lastrowid)
                
            for item in sheet_ac_data:
                cur.execute('''
                    INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, campus_id, notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (student_ids[item[0]], 2026, item[1], item[2], item[3], item[4]))
                
            for item in sheet_fee_data:
                cur.execute('''
                    INSERT INTO fees (student_id, month, year, paid_amount, date_paid, campus_id, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (student_ids[item[0]], item[1], item[2], item[3], item[4], item[5], item[6]))
            conn.commit()
            
        print(f"  Processed Sheet '{sheetname}' -> Class '{mapped_class}': {len(sheet_students_data)} students, {len(sheet_fee_data)} fee transactions.")
        
    conn.commit()
    conn.close()
    
    print("\n========================================================")
    print("         28 CAMPUS IMPORT COMPLETED SUCCESSFULLY!        ")
    print("========================================================")
    print(f"Total Students Imported: {total_students_imported}")
    print(f"   Active Students: {total_active}")
    print(f"   Withdrawn (Red/WD) Students: {total_withdrawn}")
    print(f"Total Monthly Fee Transactions: {total_fee_records} (Total: Rs. {total_fee_amount:,})")
    print(f"Total Annual / Stationery Charges: {total_ac_records} (Total: Rs. {total_ac_amount:,})")
    print("========================================================\n")



EXCEL_44_PATH = os.path.join(BASE_DIR, 'Fee Record 44 campus.xlsx')

MONTH_44_MAP = {
    'mar': ('March', 3, 2026),
    'march': ('March', 3, 2026),
    'apr': ('April', 4, 2026),
    'april': ('April', 4, 2026),
    'may': ('May', 5, 2026),
    'jun': ('June', 6, 2026),
    'june': ('June', 6, 2026),
    'jul': ('July', 7, 2026),
    'july': ('July', 7, 2026),
    'aug': ('August', 8, 2026),
    'august': ('August', 8, 2026),
    'sep': ('September', 9, 2026),
    'sept': ('September', 9, 2026),
    'september': ('September', 9, 2026),
    'oct': ('October', 10, 2026),
    'october': ('October', 10, 2026),
    'nov': ('November', 11, 2026),
    'november': ('November', 11, 2026),
    'dec': ('December', 12, 2026),
    'december': ('December', 12, 2026),
    'jan': ('January', 1, 2027),
    'january': ('January', 1, 2027),
    'feb': ('February', 2, 2027),
    'february': ('February', 2, 2027)
}

CLASS_44_MAP = {
    'p.g': 'PG',
    'pg': 'PG',
    'nursery': 'Nursery',
    'prep': 'Prep',
    'one': 'One',
    'two': 'Two',
    'three': 'Three',
    'four': 'Four',
    'five': 'Five',
    'six': 'Six',
    'seven 1': 'Seven',
    'seven': 'Seven',
    'pre 9': 'Eight',
    'pre 9th': 'Eight',
    '9th': 'Nine',
    '10th old': 'Graduate',
    'ten new': 'Ten'
}


def import_44_campus(file_path=None):
    init_db()
    
    target_excel = file_path or EXCEL_44_PATH
    if not os.path.exists(target_excel):
        print(f"Error: {target_excel} not found!")
        return
        
    print(f"\n========================================================")
    print(f"       IMPORTING 44 CAMPUS FEE RECORD                    ")
    print(f"========================================================")
    print(f"Loading Excel file: {target_excel}...")
    wb = openpyxl.load_workbook(target_excel, data_only=True)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # 1. Identify 44 Campus ID
    campus_row = cur.execute("SELECT id, name FROM campuses WHERE code = '44_2l' OR code = '44_campus' OR name LIKE '%44%2l%' OR name LIKE '%44%' ORDER BY id ASC LIMIT 1").fetchone()
    if campus_row:
        campus_id = campus_row['id']
        campus_name = campus_row['name']
    else:
        cur.execute("INSERT INTO campuses (name, code) VALUES ('44_2l campus', '44_2l')")
        conn.commit()
        campus_id = cur.lastrowid
        campus_name = '44_2l campus'
        
    print(f"Target Campus: {campus_name} (ID: {campus_id})")
    
    # 2. Clear previous records for 44 Campus to prevent duplicates
    cur.execute("DELETE FROM fees WHERE campus_id = ?", (campus_id,))
    cur.execute("DELETE FROM annual_charges_payments WHERE campus_id = ?", (campus_id,))
    cur.execute("DELETE FROM students WHERE campus_id = ?", (campus_id,))
    conn.commit()
    print("Previous records for 44 Campus cleared successfully.\n")
    
    total_students_imported = 0
    total_active = 0
    total_withdrawn = 0
    total_fee_records = 0
    total_fee_amount = 0
    total_books_records = 0
    total_books_amount = 0
    total_ac_records = 0
    total_ac_amount = 0
    
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        
        # Dynamically locate header row
        header_r = None
        for r in range(1, min(8, ws.max_row + 1)):
            vals = [str(ws.cell(r, c).value).strip().lower() for c in range(1, min(20, ws.max_column + 1)) if ws.cell(r, c).value is not None]
            if any('name' in v for v in vals):
                header_r = r
                break
                
        if not header_r:
            continue
            
        headers = [str(ws.cell(header_r, c).value).strip() if ws.cell(header_r, c).value is not None else '' for c in range(1, ws.max_column + 1)]
        while headers and not headers[-1]:
            headers.pop()
            
        mapped_class = CLASS_44_MAP.get(sheetname.lower().strip(), sheetname.strip())
        
        # Identify columns
        gross_col = None
        ac_col = None
        sp_col = None
        statio_col = None
        book_col = None
        month_cols = []
        
        for c in range(4, len(headers) + 1):
            h = headers[c-1].lower().strip()
            if 'gross' in h or 'basic' in h:
                gross_col = c
            elif h in ('a/c', 'a.c', 'ac', 'annual', 'annual charges'):
                ac_col = c
            elif 's .pack' in h or 's.pack' in h or 'spack' in h or 'summer' in h:
                sp_col = c
            elif 'statio' in h:
                statio_col = c
            elif 'book' in h or 'n.b' in h or 'n.books' in h:
                book_col = c
            else:
                for mk, mtuple in MONTH_44_MAP.items():
                    if h.startswith(mk):
                        month_cols.append((c, headers[c-1], mtuple))
                        break
                        
        sheet_students_data = []
        sheet_ac_data = []
        sheet_fee_data = []
        
        for r in range(header_r + 1, ws.max_row + 1):
            name_val = ws.cell(r, 2).value
            if not name_val or str(name_val).strip().lower() in ('none', 'nan', '', 'total', 'grand total', 'name', 'sr. #', 'sr#', 'sr. no', 'sr no'):
                continue
                
            name = str(name_val).strip()
            father_val = ws.cell(r, 3).value
            father = str(father_val).strip() if father_val and str(father_val).strip().lower() not in ('none', 'nan') else ''
            
            is_red = is_row_red(ws, r)
            has_wd_text = False
            
            # 1. Parse Monthly Basic Fee (Gross Fee column, fallback to standard 2200)
            m_fee = 0.0
            if gross_col:
                gv = ws.cell(r, gross_col).value
                if isinstance(gv, (int, float)):
                    m_fee = float(gv)
                elif gv is not None:
                    m = re.search(r'\d+', str(gv))
                    if m:
                        m_fee = float(m.group())
            if m_fee == 0.0:
                m_fee = 2200.0 # Standard base fee for 44 Campus
                
            # 2. Parse Annual Charges
            ac_val = ws.cell(r, ac_col).value if ac_col else None
            student_ac = 0.0
            ac_paid = 0.0
            ac_note = ''
            if ac_val is not None:
                if isinstance(ac_val, (int, float)):
                    student_ac = float(ac_val)
                    ac_paid = float(ac_val)
                    ac_note = 'Paid'
                else:
                    ac_str = str(ac_val).strip().lower()
                    if any(w in ac_str for w in ['w/d', 'withdraw', 'with drawl']):
                        has_wd_text = True
                        student_ac = 0.0
                        ac_paid = 0.0
                    elif ac_str in ('-', ' -', 'none', 'nan', ''):
                        student_ac = 2200.0
                        ac_paid = 0.0
                    else:
                        m = re.search(r'\d+', ac_str)
                        if m:
                            student_ac = float(m.group())
                            ac_paid = float(m.group())
                            ac_note = 'Paid'
            else:
                student_ac = 2200.0
                
            # 3. Parse Monthly Fee Payments
            student_fees = []
            first_active_month = None
            for c, h_name, (mname, mnum, myear) in month_cols:
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != '':
                    if isinstance(v, (int, float)):
                        val_num = float(v)
                        if val_num > 0:
                            if first_active_month is None:
                                first_active_month = (mnum, myear)
                            student_fees.append((mname, myear, val_num, f'{myear}-{mnum:02d}-01', 'Imported from Excel'))
                    else:
                        v_str = str(v).strip().lower()
                        if any(w in v_str for w in ['w/d', 'withdraw', 'with drawl', 'transfer', 'move']):
                            has_wd_text = True
                        elif v_str in ('-', ' -', 'none', 'nan', 'free', 'sick'):
                            pass
                        else:
                            m = re.search(r'\d+', v_str)
                            if m:
                                amt = float(m.group())
                                if amt > 0:
                                    if first_active_month is None:
                                        first_active_month = (mnum, myear)
                                    student_fees.append((mname, myear, amt, f'{myear}-{mnum:02d}-01', f'Imported from Excel ({v})'))
                                    
            status = 'withdrawn' if (is_red or has_wd_text) else 'active'
            if status == 'withdrawn':
                total_withdrawn += 1
            else:
                total_active += 1
                
            start_month = first_active_month[0] if first_active_month else 3
            start_year = first_active_month[1] if first_active_month else 2026
            
            student_idx = len(sheet_students_data)
            sheet_students_data.append((
                name, father, mapped_class, m_fee, student_ac, 0.0,
                start_month, start_year, campus_id, status
            ))
            total_students_imported += 1
            
            # Annual Charges Payment
            if ac_paid > 0:
                sheet_ac_data.append((student_idx, ac_paid, '2026-03-01', campus_id, f'Annual Charges: {ac_note}'))
                total_ac_records += 1
                total_ac_amount += ac_paid
                
            # Summer Pack Payment
            if sp_col:
                sp_val = ws.cell(r, sp_col).value
                if sp_val is not None:
                    if isinstance(sp_val, (int, float)) and float(sp_val) > 0:
                        sheet_ac_data.append((student_idx, float(sp_val), '2026-06-01', campus_id, 'Summer Pack: S.Pack Payment'))
                        total_ac_records += 1
                        total_ac_amount += float(sp_val)
                    elif isinstance(sp_val, str):
                        m = re.search(r'\d+', sp_val)
                        if m and float(m.group()) > 0:
                            amt = float(m.group())
                            sheet_ac_data.append((student_idx, amt, '2026-06-01', campus_id, 'Summer Pack: S.Pack Payment'))
                            total_ac_records += 1
                            total_ac_amount += amt
                            
            # Stationery Payment (Playgroup, Nursery, Prep)
            if statio_col:
                st_val = ws.cell(r, statio_col).value
                if st_val is not None:
                    if isinstance(st_val, (int, float)) and float(st_val) > 0:
                        sheet_ac_data.append((student_idx, float(st_val), '2026-03-01', campus_id, 'Stationery Payment'))
                        total_ac_records += 1
                        total_ac_amount += float(st_val)
                    elif isinstance(st_val, str):
                        m = re.search(r'\d+', st_val)
                        if m and float(m.group()) > 0:
                            amt = float(m.group())
                            sheet_ac_data.append((student_idx, amt, '2026-03-01', campus_id, 'Stationery Payment'))
                            total_ac_records += 1
                            total_ac_amount += amt
                            
            # Books Payment
            if book_col:
                bk_val = ws.cell(r, book_col).value
                if bk_val is not None:
                    if isinstance(bk_val, (int, float)) and float(bk_val) > 0:
                        sheet_ac_data.append((student_idx, float(bk_val), '2026-03-01', campus_id, 'Books / Syllabus Payment'))
                        total_books_records += 1
                        total_books_amount += float(bk_val)
                    elif isinstance(bk_val, str):
                        m = re.search(r'\d+', bk_val)
                        if m and float(m.group()) > 0:
                            amt = float(m.group())
                            sheet_ac_data.append((student_idx, amt, '2026-03-01', campus_id, 'Books / Syllabus Payment'))
                            total_books_records += 1
                            total_books_amount += amt

                            
            # Monthly Fee Payments
            for mname, myear, amt, date_paid, note in student_fees:
                sheet_fee_data.append((student_idx, mname, myear, amt, date_paid, campus_id, note))
                total_fee_records += 1
                total_fee_amount += amt
                
        # Batch Insert for Sheet
        if is_postgres():
            from psycopg2.extras import execute_values
            raw_conn = conn._conn
            raw_cur = raw_conn.cursor()
            
            if sheet_students_data:
                res = execute_values(
                    raw_cur,
                    """INSERT INTO students (name, father_name, class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, campus_id, status)
                       VALUES %s RETURNING id""",
                    sheet_students_data,
                    fetch=True
                )
                student_ids = [row[0] for row in res]
                
                if sheet_ac_data:
                    ac_to_insert = [
                        (student_ids[item[0]], 2026, item[1], item[2], item[3], item[4])
                        for item in sheet_ac_data
                    ]
                    execute_values(
                        raw_cur,
                        """INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, campus_id, notes)
                           VALUES %s""",
                        ac_to_insert
                    )
                    
                if sheet_fee_data:
                    fees_to_insert = [
                        (student_ids[item[0]], item[1], item[2], item[3], item[4], item[5], item[6])
                        for item in sheet_fee_data
                    ]
                    execute_values(
                        raw_cur,
                        """INSERT INTO fees (student_id, month, year, paid_amount, date_paid, campus_id, notes)
                           VALUES %s""",
                        fees_to_insert
                    )
            raw_conn.commit()
        else:
            student_ids = []
            for s_row in sheet_students_data:
                cur.execute('''
                    INSERT INTO students (name, father_name, class, monthly_fee, annual_charges, opening_arrears, start_month, start_year, campus_id, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', s_row)
                student_ids.append(cur.lastrowid)
                
            for item in sheet_ac_data:
                cur.execute('''
                    INSERT INTO annual_charges_payments (student_id, year, paid_amount, date_paid, campus_id, notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (student_ids[item[0]], 2026, item[1], item[2], item[3], item[4]))
                
            for item in sheet_fee_data:
                cur.execute('''
                    INSERT INTO fees (student_id, month, year, paid_amount, date_paid, campus_id, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (student_ids[item[0]], item[1], item[2], item[3], item[4], item[5], item[6]))
            conn.commit()
            
        print(f"  Processed Sheet '{sheetname}' -> Class '{mapped_class}': {len(sheet_students_data)} students, {len(sheet_fee_data)} fee/book transactions.")
        
    conn.commit()
    conn.close()
    
    print("\n========================================================")
    print("         44 CAMPUS IMPORT COMPLETED SUCCESSFULLY!        ")
    print("========================================================")
    print(f"Total Students Imported: {total_students_imported}")
    print(f"   Active Students: {total_active}")
    print(f"   Withdrawn (Red/WD) Students: {total_withdrawn}")
    print(f"Total Monthly Tuition Transactions: {total_fee_records} (Total: Rs. {total_fee_amount:,})")
    print(f"Total Books Payment Transactions: {total_books_records} (Total: Rs. {total_books_amount:,})")
    print(f"Total Annual / Summer Pack / Stationery Charges: {total_ac_records} (Total: Rs. {total_ac_amount:,})")
    print("========================================================\n")


if __name__ == '__main__':
    import sys
    arg = sys.argv[1].lower() if len(sys.argv) > 1 else 'main'
    
    if arg in ('44', '44_campus', 'campus44', '44_2l'):
        import_44_campus()
    elif arg in ('28', '28_campus', 'campus28'):
        import_28_campus()
    elif arg in ('all', 'both', 'three'):
        import_main_campus()
        import_28_campus()
        import_44_campus()
    else:
        import_main_campus()


